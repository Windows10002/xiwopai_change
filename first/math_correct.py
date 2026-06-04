from __future__ import annotations

import os
import re
import json
import logging
import math
import textwrap
import base64
import sys
import subprocess
import ast
import operator
from fractions import Fraction
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from http import HTTPStatus
try:
    import dashscope
except ImportError:
    dashscope = None  # type: ignore
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ================= 1. 环境初始化 =================
if sys.platform == 'win32':
    try:
        subprocess.run(['chcp', '65001'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

init(autoreset=True)
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

# ================= 2. 配置区 =================
if dashscope is not None:
    dashscope.api_key = os.getenv("DASHSCOPE_API_KEY", "")
MATH_FOLDER = "img_math"
OUTPUT_EXCEL = "数学作业批改报告.xlsx"
OUTPUT_TXT = "数学作业批改报告.txt"

# ================= 3. 核心 Prompt =================
MATH_PROMPT = """
你是资深小学/初中数学阅卷教师。请公平、审慎批改图片中的手写作答：既要防止明显错误被放过，更要防止「推理与答案实质正确却被判错」尤其是 OCR 把下标、乘号、涂改笔画误读时。

【识别要求】
1. 必须先完整找出图片中的所有题号/横式/竖式/应用题，逐题输出，禁止遗漏；不要只输出写了答案的题。
2. 对每个题号先判断“是否有有效作答”。仅当答题区完全没有任何手写数字时，才判【未作答】。只要该题区域内识别到任意手写数字（含只写最终结果、竖式末行、漏写小数点的错位数），一律视为已作答，不得判【未作答】、不得称「串题/无关」。只有完全空白、仅题号、仅无关涂画、答案被划掉且无替代、无任何数字时，才判【未作答】。
2b. **questions 数组必须覆盖卷面上每一道可辨认题目（含横式/竖式/脱式/应用题），题量须与卷面一致，禁止只输出前几题或省略中间题号；若一页超过 20 题仍须全部输出，不得因长度自行截断。**
2c. **仅把原题横式/算式照抄一遍、没有「=」后的有效变形或计算、或「=」两侧只是重复题干而无新数值，一律判【错误】或【未作答】，process_score、result_score、structure_score 全部为 0。**
2d. **划改识别（极其重要）**：凡被手写横线、叉号、涂改液、橡皮擦淡影明显划掉的一行或一整段演算，一律视为作废，不得写入 student_work，不得在 reason 中引用、摘抄或点评；只保留划改之后最终仍清晰可见的有效书写。若划改后另起一行重写，以重写行为准。
2f. **student_work 与 reason 必须可核对（极其重要）**：`student_work` 只能抄写卷面上该题号下**真实可见**的手写脱式/竖式，看不清的用「疑似」标注，**禁止**编造卷面没有的数字或通分步骤。`reason` 里描述「学生先算…再算…」时，每一步必须能在 `student_work` 中找到对应片段；若卷面无有效作答，`student_work` 与 `student_answer` 必须为空并判【未作答】，**禁止**在 reason 里虚构「学生算出 1又1/12」等过程。
2e. **首行「=」右侧优先（极其重要）**：若紧接「=」的右侧已出现清晰、连贯的有效脱式（例如「30×42」或「30 × 42」），而「=」左侧仅有涂抹、杂乱笔画、被横线划掉的区域或无法辨认的乱迹，**禁止**把左侧乱迹误读成算式片段再与右侧拼接（典型误读：把乱迹+「30×42」读成「46-30×42」）。此时应认定学生已先完成括号内运算得到 30，再与 42 相乘；student_work 只收录「=」右侧起的有效链。
3. 严禁把下一题/上一题的作答挪到本道题。每题只能读取该题号右侧、下方且空间归属明确的书写；若本道题答题区无任何手写数字，才判【未作答】——不得因下一题有步骤就判本题未作答。
4. 学生明显涂改、划掉、叉掉、擦除残影、重写前的旧过程，一律视为无效痕迹，不能写入 student_work，不能参与过程分和结果分；只批改最后未被划掉的有效作答。reason 中也不要把被划掉的错误数字/错误步骤当成错误原因。
5. 遇到一页中有多题时，必须按从上到下、从左到右顺序逐题编号；看不清但能确认题目存在时也要输出，并在 reason 中说明“题目/作答疑似模糊”。
6. OCR 不确定时，在 reason 中说明「疑似识别为…」，但仍要按可见内容批改，不能因为不确定就默认正确。
6b. **印刷体与手写易混（极其重要）**：`+` 与 `÷`、`×` 与 `x`、小数点与乘号、带分数「整数 空格 分数」与「整数乘分数」要对照印刷题干整体判断；**若印刷体明确为 (a-b)×c、(a+b)÷d 等括号结构，学生先算括号再乘除且数值链成立，不得因 expr 字段抄错运算符而改判为错。**
6c. **expr 字段**必须与卷面印刷题干一致；若题干含括号而你写的 expr 丢了括号或把 × 写成 +，应先按图片修正 expr 再批改，避免后续「标准结果」与卷面不符。
6d. **除号与分数（极其重要）**：题干或脱式中出现「÷ a/b」（a、b 为整数）时，必须理解为「÷ (a/b)」，即除以整个分数 **a/b**，**禁止**理解为「(÷ a) ÷ b」或从左到右拆成 `/a/b` 导致数值错误。同级 **× 与 ÷** 仍按课标从左到右依次计算；**禁止**先做后面的乘法再做前面的除法。
6d2. **带分数与「+1/3」易混（极其重要）**：印刷体在「-1/6 - 1/4 + ___」这类括号末项若是 **带分数 1 1/3**，整数与分数之间常有空隙，切勿误读为单独的 **+1/3**；expr 必须按卷面带分数抄写（写作 `1 1/3`），否则整题标准值会错。
6d2b. **分母 7 与 4 易混（极其重要）**：印刷体带分数如 **3 1/7、-3 1/7** 的分母「7」常被误读成「4」。同一式中若另有 **5 1/4** 等分母为 4 的项，须逐项对照卷面竖式分母，**禁止**把括号内的 1/7 抄成 1/4；expr 中每一项带分数的分母必须与图片一致。
6d3. **带分数与假分数书写等价（极其重要）**：学生在括号内写 **1 1/3** 或与它数值等价的 **4/3、16/12** 等通分形式，只要等号链每步成立、结果与按题干重算一致，**不得**声称「把 1 1/3 写成假分数所以通分错」；**禁止**把真分数 **11/12** 误写成带分数「4又 1/12」等离奇形式来否定学生；更不得把学生已算对的 **144/11** 说成与标准答案不一致。
6e. **expected_answer** 必须与按 expr 在实数范围内重算的结果一致（允许带分数、小数、最简分数等书写等价）；若你写的 expected_answer 与重算不符，以重算为准，不得据此把学生判错。
识别图片中的数学题目，把所有分数、除法符号转换成普通文本格式，比如把\frac{1}{3}写成1/3，把\div写成÷，不要输出任何 LaTeX 代码或转义字符。
这样能大幅减少乱码概率。
【计算核验规则】
1. 对每道计算题必须独立重算标准结果，再核对学生过程和最终答案；先验算题干原式，再验算每一个等号右侧是否由左侧正确推出。
2. “按步骤给分”只奖励错误发生前的正确步骤：如果第一步就抄题、变形、运算顺序、括号处理、符号或等号关系错误，process_score 必须为 0；不能因为后面继续算对了错误式子而给过程分。
2b. **脱式首行省略（极其重要）**：若学生从中间某步写起（如仅「=976÷8=…」而前面「(608+368)÷8」在卷面上行已写清），只要该步数值与题干在该阶段应得值一致，**不得**把「第一个等号」误判为从原式第一步就错；仅最后一步算术抄错时，须保留过程分（通常 4～6），result_score=0。
3. 对「纯数字四则/脱式」：中间过程出现等号不成立、运算对象抄错、括号/符号/顺序错等，即使最终答案碰巧与正确答案相同，也判【错误】。对「方程、配方、因式分解、开平方得±、多根」等含未知数的题：不得以「相邻两步之间没有用纯数字验算」为由判整条错误；只要配方/开方/求根思路正确且求得的根代入能满足方程，应判【正确】或【过程不规范】，不得判【错误】。
3b. **末步算术错误（极其重要）**：若前几步括号与变形均正确，仅最后一步加减乘除算错，使 student_answer 与 expected_answer 数值不一致，**禁止**判【正确】；必须判【错误】，result_score=0，process_score 通常给 4～6 表示「前半过程正确」。
3c. **仅符号相反（极其重要）**：括号内通分、除法变乘法等前几步均正确，仅最后一步把「÷(-1/12)」写成「×12」等导致结果为正、标准值为负（或反之），**必须**保留过程分（通常 4～6），result_score=0，**禁止**套用「第一步已从原式错误变形」清零；reason 应点明「符号错误」。
3d. **多步小数与带分数混合**：若学生将可精确抵消的项（如 -5.25 与 +5¼）先行抵消，再把余下的带分数与整数近似合并书写（如写作「+6」），只要脱式等号链在**其书写下自洽**且最终分式与按题干严格通分结果相差不超过 **1/4**（初中练习常见取整/近似容差），判【正确】或【过程不规范】，不得仅因与冗长通分式数值相差极小而判【错误】。
3d2. **跳步简便运算（极其重要）**：学生首行未逐字抄全原式，但已将若干项**合理合并**（如抵消小数与带分数、先合并同分母分数），只要**每一个等号两侧数值成立**且**最终结果与按题干重算一致**，必须判【正确】或【过程不规范】，**禁止**套用「第一步已从原式错误变形」清零，也**禁止**在订正提示中写「末步与标准答案不一致」类自相矛盾表述。
3e. **先化简再求值**：若化简式与代入后的数值答案与 expected_answer 一致，即使 reason 中挑剔中间某行笔误，只要 student_answer 正确，必须判【正确】或【过程不规范】，不得判【错误】。
4. 如果过程描述为“过程正确、每步成立、结果正确”，status 必须是【正确】或【过程不规范】，不得同时判为【错误】；除非明确指出哪一步具体错误。
4b. **根式/绝对值/实数混合运算（极其重要）**：若 reason 已写「结果正确」「合并后得」等肯定语，status 必须为【正确】或【过程不规范】；**禁止**再写「订正提示」「末步与标准答案不一致」；\(2+\sqrt{2}\) 与 \(\sqrt{2}+2\) 等等价写法一律视为正确。
5. 示例：(52-7)×6 写成 =52-(7×6) 是第一步错误，process_score=0，标准步骤为 (52-7)×6=45×6=270。
6. 示例：979÷(144-133) 如果本道题无有效书写，不得把下一题的 (16+34)×30 识别成本题作答；应判【未作答】。
7. 示例：21×(60-55)=21×5=105 每一步都成立，必须判【正确】，不能因模板化语句写成错误。
8. 示例：22×(16+21) 中被涂改/划掉的错误过程要忽略，只看最后保留的有效过程；如果最后有效过程正确，应判正确。再例如学生先写了错误数字又划掉，后面改成正确答案，划掉内容不得出现在 student_work 和 reason 中。
8b. 示例：(46-16)×42 卷面若在第一个「=」左侧有重涂/涂改，只要「=」后有效书写为「30×42 = 1260」且每步数值成立，必须按「先算括号得 30 再乘 42」理解并优先判【正确】；**除非**在未被划掉的区域真的写出了「46-30×42」这类整段式子，否则不得在 reason 中声称学生写成了「46-30×42」或「违反运算顺序」。
9. 常见题型都要严查：整数/小数/分数四则混合运算、脱式计算、简便计算、竖式计算、解方程、比例、百分数、单位换算、应用题列式、几何周长面积体积、带余除法。

【百分数运算（极其重要，口算/脱式通用）】
1. **50% 表示 0.5 或 50/100，绝不是整数 50**；30% 是 0.3，25% 是 0.25，110% 是 1.1。验算与 expected_answer 必须先换算再算，**禁止**去掉百分号后当普通整数。
2. 示例：0.55-50% = 0.55-0.5 = **0.05**（不是 0.55-50=-49.45）；1-25% = 1-0.25 = **0.75**（不是 -24）；5/8-50% = 5/8-0.5 = **1/8**。
3. 示例：0.8÷[(5/8-50%)÷5/8] 中 50% 必须先化为 0.5，括号内 5/8-50% = 1/8，再按运算顺序得 **4**。
4. 含 % 的乘除：66÷33% = 66÷0.33 = 200；30×75% = 22.5；24%÷48% = 0.5；35%×0.2 = 0.07；36%+47.1% = 0.831；80%+110% = 1.9；13%+0.77 = 0.9；28%÷0.56 = 0.5；30%×30% = 0.09。
5. **每一个**百分数项都必须先 ÷100 再参与运算，禁止只转换式中部分 %、或把式首的 35% 当成 35。
6. expr 字段须保留 % 符号（如 `0.55-50%`），expected_answer 填换算后的数值结果。

10. 允许合理等价形式：分数等价、约分后相等、乘除互化、交换律/结合律/分配律正确使用；但“等价”必须每步成立。
11. 竖式题要核对每一位、进位/退位/小数点位置；应用题要核对数量关系、列式、单位和答句；方程题要核对移项/合并/两边同乘除/检验。

【评分权重，满分 10 分】
- process_score 过程分 6 分：列式/方法 2 分，关键步骤与运算顺序 2 分，每步计算细节 2 分。
- result_score 结果分 3 分：数值正确 2 分，最终形式/化简/单位正确 1 分。
- structure_score 规范分 1 分：书写格式、答句、单位、约分、竖式对齐等规范。解方程时 x1/x2 与 x₁/x₂、未单独写「解集」或检验步骤，**不**视为格式错误，不得因此把【正确】降为【过程不规范】或扣光 structure_score。
- 过程正确但结果错：process_score 通常给 4-6，result_score 必须为 0，structure_score 按书写给 0-1，status 为【错误】。
- 过程基本对但有轻微算术失误：process_score 通常给 3-5，result_score 为 0，status 为【错误】。
- 「纯算术」结果正确但脱式中等号链不成立/抄题错：status 必须为【错误】；如果第一步就错，process_score=0；如果中途才错，process_score 只给错误前已正确步骤，通常不超过 3；result_score 最多给 1-2，不能因答案碰巧相同给满分。含未知数的方程/配方/开方题不适用「逐段纯数字验等号」的严苛标准，见上【计算核验规则】第 3 条。
- 只有答案无过程：若题目要求过程/脱式/竖式/应用题，process_score 不超过 2；若答案正确可给 result_score，status 通常为【过程不规范】。
- 未作答：三项均为 0，status 为【未作答】。
- 被涂改/划掉的旧过程：不计入 student_work，不给分也不扣分，只按最终保留的有效作答评分。
- 错误或未作答题目必须在 reason 中给出注解，并写出完整正确步骤及答案。

【小数计算 / 小数乘法 / 积的近似数（极其重要）】
1. expected_answer、reason、【正确步骤】【正确答案】及订正展示全程只用小数（如 1.3、1000、0.22），禁止分数、带分数、通分步骤。
2. 积的近似数：先算精确小数乘积，再按题意四舍五入保留指定位数；末位与标准近似值一致即【正确】，不得因中间写了未四舍五入的精确积而判错。expected_answer 必须填写四舍五入后的最终得数，不得填精确积。
3. 若 reason 已写明「符合要求/符合题意」且学生末位为规范近似值，必须判【正确】并给满分，禁止再写「与标准答案不一致」「订正提示」等自相矛盾语句。
4. 错误题 reason 精简为一句：【主因标签】短说明 + 正确步骤：横式=结果，答案：小数。主因标签只能选一个：【计算结果错误】【运算顺序错误】【小数点位置错误】【取值近似错误】。禁止混用多种主标签，禁止写「过程正确、结果错误」与「小数点位置错误」同时出现。
5. 禁止冗长套话：「说明：」「按题干重算」「与当前作答不一致」「末步结果与标准答案不一致」「订正提示」「结果分记 0 分」等一律不写。

【订正说明 reason 的书写规范（面向学生展示）】
1. 只用「老师批改学生作业」的口吻，短句、条理清楚；禁止出现「模型/AI/大模型/OCR/识别器/置信度/猜测/分析过程」等字样，禁止自我辩解或复述你的内心推理。
2. **正确**或**过程不规范**题：用一两句肯定写对之处即可，不必写长篇「解析」。
3. **错误**题（分数/方程/化简等非常规小数卷）：可用四段模板【错误原因】【正确步骤】【正确答案】【薄弱点】；**小数计算、小数乘法、积的近似数**一律用上一节「一句式」格式，不用四段模板。
4. **正确**或**过程不规范**题：一两句肯定即可；不要写四段模板。
5. 化简/整式题：按同类项分组说明（如含 ab、a³b²、常数项分别合并），不要漏项；不要凭直觉给答案。
6. 纯算术题若学生最终结果与标准答案不同，必须判【错误】；若过程不成立即使结果碰巧相同也必须判【错误】。

【状态标准】
- 【正确】：过程、结果、规范都正确，总分通常 9-10。
- 【过程不规范】：核心方法、变形与最终数值/根正确，仅有轻微跳步、未写检验/解集集合形式、下标写成 x1/x2 与 x₁/x₂ 混用、书写略乱等，总分通常 7-9；不得仅因此类格式问题判【错误】。
- 【错误】：答案错、关键变形错、抄题错、符号错、计算错、逻辑错，或「纯算术题」中结果碰巧对但过程不成立。
- 【未作答】：题目存在但没有有效作答。

【输出要求】
1. 输出必须为标准 JSON，禁止 Markdown 代码块或额外文字。
2. 算式统一用纯文本：分数写为 1/3，带分数写为 1 1/3，指数写为 x^2。

【JSON 格式】
{
  "overall_result": "全部正确/部分正确/全部错误/未作答",
  "questions": [
    {
      "id": 1,
      "expr": "题目原式",
      "student_work": "学生可见作答/过程；未作答则为空字符串",
      "expected_answer": "标准结果",
      "student_answer": "学生最终答案；没有则为空字符串",
      "status": "正确/过程不规范/错误/未作答",
      "process_score": 0,
      "result_score": 0,
      "structure_score": 0,
      "total_score": 0,
      "reason": "大白话说明：哪里对、哪里错、标准算法是什么"
    }
  ],
  "personal_comment": "约 80～120 字：先总评对错分布与运算能力；再各用一两句点评书写整洁度、审题细心程度、验算习惯；最后点出最值得巩固的一条建议。语气亲切、具体、可执行。",
  "weak_points": ["知识点 1", "知识点 2"]
}
"""


# ================= 4. 核心工具函数 =================

def extract_json_safe(text: str) -> str:
    block = re.search(r'```json\s*(.*?)\s*```', text, re.DOTALL)
    if block: return block.group(1)
    start, end = text.find('{'), text.rfind('}')
    if start != -1 and end != -1 and end > start:
        return text[start: end + 1]
    raise ValueError("无法定位 JSON")


def prettify_math(text: str) -> str:
    if not text: return ""
    sup_map = str.maketrans("0123456789-", "⁰¹²³⁴⁵⁶⁷⁸⁹⁻")
    text = re.sub(r'\^([\d-]+)', lambda m: m.group(1).translate(sup_map), text)
    text = text.replace('/', '⁄')
    return text


def get_mime_type(path: str) -> str:
    ext = Path(path).suffix.lower()
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "bmp": "image/bmp",
            "webp": "image/webp"}.get(ext, "image/jpeg")


def clamp_score(value, min_value=0, max_value=10):
    try:
        value = float(value)
    except (TypeError, ValueError):
        value = 0
    return max(min_value, min(max_value, value))


_ALLOWED_AST_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def _safe_eval_ast(node):
    if isinstance(node, ast.Expression):
        return _safe_eval_ast(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return Fraction(str(node.value))
    if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_AST_OPS:
        left = _safe_eval_ast(node.left)
        right = _safe_eval_ast(node.right)
        if right == 0 and isinstance(node.op, (ast.Div, ast.FloorDiv, ast.Mod)):
            raise ZeroDivisionError
        return _ALLOWED_AST_OPS[type(node.op)](left, right)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_AST_OPS:
        return _ALLOWED_AST_OPS[type(node.op)](_safe_eval_ast(node.operand))
    raise ValueError("unsupported expression")


def preprocess_decimal_ocr_text(text: str) -> str:
    """OCR 后处理：统一小数点写法，减少漏识/误识为小数点前的噪点。"""
    if not text:
        return ""
    t = str(text)
    t = t.replace("．", ".").replace("·", ".")
    t = re.sub(r"(\d)\s*[,，]\s*(\d)", r"\1.\2", t)
    t = re.sub(r"(\d)\s*\.\s*(\d)", r"\1.\2", t)
    t = re.sub(r"(\d{2,})\s+00(?!\d)", r"\1.00", t)
    return t


_UNICODE_VULGAR_FRACTIONS = {
    "½": "1/2",
    "⅓": "1/3",
    "⅔": "2/3",
    "¼": "1/4",
    "¾": "3/4",
    "⅕": "1/5",
    "⅖": "2/5",
    "⅗": "3/5",
    "⅘": "4/5",
    "⅙": "1/6",
    "⅚": "5/6",
    "⅛": "1/8",
    "⅜": "3/8",
    "⅝": "5/8",
    "⅞": "7/8",
}


def preprocess_chinese_math_text(text: str) -> str:
    """手写带分数、题头标签等中文书写，须在去中文前规范化。"""
    if not text:
        return ""
    t = preprocess_decimal_ocr_text(str(text))
    # 1⅓、2½ 等紧贴写法 → 带分数（须在单字符替换之前）
    t = re.sub(
        r"(-?\d+)([½⅓⅔¼¾⅕⅖⅗⅘⅙⅚⅛⅜⅝⅞])",
        lambda m: f"{m.group(1)} {_UNICODE_VULGAR_FRACTIONS.get(m.group(2), m.group(2))}",
        t,
    )
    for ch, repl in _UNICODE_VULGAR_FRACTIONS.items():
        t = t.replace(ch, repl)
    t = re.sub(r"(\d+)\s*[又帯带]\s*(\d+)\s*/\s*(\d+)", r"\1 \2/\3", t)
    t = re.sub(r"(\d+)\s*[又帯带]\s*(\d+)\s*[/／]\s*(\d+)", r"\1 \2/\3", t)
    # OCR/模型常把「1 1/3」拆成「1 1÷3」「1 1/3」前的「1 1 ÷ 3」
    t = re.sub(r"(?<![\d/.])(\d+)\s+1\s*[÷/]\s*3(?!\d)", r"\1 1/3", t)
    return t


_PERCENT_LITERAL_RE = re.compile(r"(?<![\d.])(\d+(?:\.\d+)?)\s*%")


def convert_percent_literals(text: str) -> str:
    """将 50%、30.5％ 等百分数转为 (50/100) 再参与求值；禁止把 % 去掉后当整数 50。"""
    if not text:
        return ""
    t = str(text).replace("％", "%")
    if "%" not in t:
        return t
    # 句首负百分数：-50% → -((50)/100)，避免与正百分数规则冲突
    t = re.sub(r"^-(\d+(?:\.\d+)?)\s*%", r"-((\1)/100)", t)
    t = _PERCENT_LITERAL_RE.sub(lambda m: f"(({m.group(1)})/100)", t)
    return t


def _percent_numbers_in_text(text: str) -> list[str]:
    raw = str(text or "").replace("％", "%")
    return _PERCENT_LITERAL_RE.findall(raw)


def is_percent_expression_context(expr: str, student_work: str = "") -> bool:
    raw = f"{expr or ''}\n{student_work or ''}"
    return "%" in raw.replace("％", "%")


def maybe_fix_expr_missing_percent(expr: str, student_work: str = "", reason: str = "") -> str:
    """模型 expr 漏写 % 时，按卷面/评语中的 50% 等补回，避免把 50% 当整数 50 验算。"""
    ex = str(expr or "").strip()
    if not ex or "%" in ex or "％" in ex:
        return ex
    hint = preprocess_chinese_math_text(f"{student_work or ''}\n{reason or ''}")
    percents = _percent_numbers_in_text(hint)
    if not percents:
        return ex
    out = ex
    for num in sorted(set(percents), key=len, reverse=True):
        if f"{num}%" in out:
            continue
        pat_after_op = re.compile(
            rf"(?<=[+\-(\[（/÷×*])\s*{re.escape(num)}(?![%\d/.])"
        )
        pat_leading = re.compile(rf"^{re.escape(num)}(?![%\d/.])")
        if pat_after_op.search(out):
            out = pat_after_op.sub(f"{num}%", out, count=1)
        elif pat_leading.search(out):
            out = pat_leading.sub(f"{num}%", out, count=1)
    return out


def normalize_math_expr(expr: str) -> str:
    """将常见手写/中文数学符号转成可安全计算的 Python 表达式。"""
    if not expr:
        return ""
    text = preprocess_chinese_math_text(str(expr))
    text = convert_percent_literals(text)
    text = text.replace("（", "(").replace("）", ")").replace("[", "(").replace("]", ")")
    # 仅把显式乘号换成 *，不要把字母 x 换成 *，否则会破坏方程、配方、因式分解等含 x 的书写。
    text = text.replace("×", "*").replace("·", "*").replace("÷", "/").replace("／", "/")
    # 禁止把「解：」「原式：」等全角冒号当成除号；仅数字比（如 3：4）才转为 /
    text = re.sub(r"(?<=\d)\s*：\s*(?=\d)", "/", text)
    text = text.replace("－", "-").replace("—", "-").replace("–", "-").replace("＋", "+")
    text = text.replace("＝", "=").replace("≈", "=").replace("≒", "=")
    text = re.sub(r"(?<=\d),(?=\d{3}(\D|$))", "", text)
    text = re.sub(r"[^0-9+\-*/().=\s]", "", text)
    return text.strip()


def rewrite_div_fraction_pairs(text: str) -> str:
    """将 ÷ 后紧跟的「整数/整数」视为除以该分数整体，避免求值成 (((a)/b)/c) 的课标外误读。"""
    if not text or "÷" not in text:
        return str(text)
    s = str(text)
    pattern = re.compile(r"÷\s*(-?\d+)\s*/\s*(-?\d+)")
    for _ in range(64):
        m = pattern.search(s)
        if not m:
            break
        repl = f"/(({m.group(1)})/({m.group(2)}))"
        s = s[: m.start()] + repl + s[m.end() :]
    return s


_MIXED_FRAC_TOKEN_RE = re.compile(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])")
# 印刷体/手写分母常见误识（如 7↔4）
_DENOM_OCR_SWAP = {
    "4": ("7", "9"),
    "7": ("4",),
    "6": ("4",),
    "9": ("4",),
}


def _student_work_has_mixed_frac(work: str, whole: str, num: str, denom: str) -> bool:
    if not work:
        return False
    w = str(work)
    esc = re.escape
    if re.search(rf"(?<![\d/.]){esc(whole)}\s+{esc(num)}\s*/\s*{esc(denom)}(?![\d/.])", w):
        return True
    return bool(re.search(rf"{esc(whole)}\s*[又帯带]?\s*{esc(num)}\s*/\s*{esc(denom)}", w))


def _mixed_frac_denominator_expr_candidates(raw: str, student_work: str) -> list[str]:
    """因带分数分母误识（如 3 1/7→3 1/4）生成题干变体；student_work 中出现的分母优先。"""
    candidates = []
    seen = set()
    for m in _MIXED_FRAC_TOKEN_RE.finditer(raw):
        whole, num, denom = m.group(1), m.group(2), m.group(3)
        alts = list(_DENOM_OCR_SWAP.get(denom, ()))
        hinted = [d for d in alts if _student_work_has_mixed_frac(student_work, whole, num, d)]
        for alt_d in hinted + [d for d in alts if d not in hinted]:
            s = raw[: m.start()] + f"{whole} {num}/{alt_d}" + raw[m.end() :]
            if s not in seen and s != raw:
                seen.add(s)
                candidates.append(s)
    return candidates


def _apply_den7_ocr_full_correction(raw: str) -> str:
    """同式中保留 5 1/4，其余带分数分母 4 一律改为 7（第 1 题等「7↔4」误识的典型整式纠错）。"""
    if not raw or not re.search(r"\d+\s+\d+\s*/\s*4", raw):
        return raw

    def repl(m: re.Match) -> str:
        whole, num, denom = m.group(1), m.group(2), m.group(3)
        if denom != "4":
            return m.group(0)
        try:
            if abs(int(whole)) == 5:
                return m.group(0)
        except ValueError:
            pass
        return f"{whole} {num}/7"

    return _MIXED_FRAC_TOKEN_RE.sub(repl, raw)


def _student_work_implies_one_and_one_third(work: str) -> bool:
    """脱式通分出现 16/12、11/12 等，说明括号内末项实为带分数 1 1/3 而非 +1/3。"""
    if not work:
        return False
    w = str(work)
    return bool(
        re.search(r"16\s*/\s*12", w)
        or re.search(r"11\s*/\s*12", w)
        or re.search(r"4\s*/\s*3", w)
        or re.search(r"1\s+1\s*/\s*3", w)
    )


def _pick_best_ocr_expr_candidate(candidates, raw, sv, student_work):
    """多个 OCR 纠错候选与学生结果数值相同时，按过程佐证与纠错幅度择优。"""
    matched = []
    for cand in candidates:
        if cand == raw:
            continue
        v = safe_eval_math_expr(cand)
        if v is not None and values_equal(v, sv):
            matched.append(cand)
    if not matched:
        return None
    if len(matched) == 1:
        return matched[0]

    work = str(student_work or "")

    def score(cand: str) -> tuple:
        den_fixes = len(re.findall(r"\d+\s+\d+\s*/\s*7", cand)) - len(re.findall(r"\d+\s+\d+\s*/\s*7", raw))
        has_one_one_third = bool(re.search(r"1\s+1\s*/\s*3", cand)) and not re.search(r"1\s+1\s*/\s*3", raw)
        work_hint = 0
        if _student_work_implies_one_and_one_third(work) and has_one_one_third:
            work_hint = 10
        if re.search(r"5\s+1\s*/\s*4", raw) and re.search(r"3\s+1\s*/\s*7", cand):
            work_hint += 5
        return (work_hint, den_fixes, -len(cand))

    return max(matched, key=score)


def _ocr_den4_to_7_expr_candidates(raw: str) -> list[str]:
    """分母 7 误识为 4：对带分数「又 1/4」尝试改为 1/7（保留合法的 5 1/4），支持多项组合替换。"""
    positions = []
    for m in _MIXED_FRAC_TOKEN_RE.finditer(raw):
        if m.group(3) != "4":
            continue
        try:
            w = int(m.group(1))
        except ValueError:
            continue
        if abs(w) == 5:
            continue
        positions.append(m)
    if not positions:
        return []
    candidates = []
    n = len(positions)
    for mask in range(1, 1 << n):
        s = raw
        for i in sorted(range(n), reverse=True):
            if not (mask & (1 << i)):
                continue
            m = positions[i]
            s = s[: m.start()] + f"{m.group(1)} {m.group(2)}/7" + s[m.end() :]
        if s != raw:
            candidates.append(s)
    return candidates


def maybe_fix_expr_from_student_numeric(expr: str, student_work: str, student_answer: str) -> str:
    """当模型把常见「1 1/3」误识成「1/3」、分母 7 误识成 4 等导致 expr 与卷面不符时，用学生作答数值反推可自洽的题干变体（仅少数安全模式）。"""
    raw = str(expr or "").strip()
    if not raw:
        return raw
    raw = raw.replace("⁄", "/")
    work = str(student_work or "")
    sv = extract_math_value(student_answer) or extract_math_value(student_work)
    if sv is None and student_work and "=" in str(student_work):
        steps_tail = split_equation_steps(str(student_work))
        if steps_tail:
            tail = steps_tail[-1]
            sv = safe_eval_math_expr(tail) or extract_math_value(tail)
    if sv is None:
        return raw
    base = safe_eval_math_expr(raw)
    if base is not None and values_equal(base, sv):
        return raw

    candidates = []
    # 印刷体「1 1/3」常被误识为 +11÷3、+11/3、1 1÷3（第 4 题等）
    for pat, repl in (
        (r"\+11\s*[÷/]\s*3(?!\d)", "+1 1/3"),
        (r"\+11\s+1\s*/\s*3", "+1 1/3"),
        (r"(?<![\d/.])1\s+1\s*[÷/]\s*3(?!\d)", "1 1/3"),
        (r"(?<![\d/.])1\s*⅓", "1 1/3"),
    ):
        s = re.sub(pat, repl, raw)
        if s != raw:
            candidates.append(s)
    # 括号内「+1/3」误少整数 1：常见于 12÷(-1/6-1/4+1 1/3) 被读成 +1/3（含 4+1/3、+1/3) 等）
    for m in re.finditer(r"\+(\s*)1(\s*/\s*)3(?!\d)", raw):
        s = raw[: m.start()] + "+1 1/3" + raw[m.end() :]
        if s != raw:
            candidates.append(s)
    s_paren = re.sub(r"\+1/3(\s*\))", r"+1 1/3\1", raw)
    if s_paren != raw:
        candidates.append(s_paren)
    candidates.extend(_mixed_frac_denominator_expr_candidates(raw, work))
    candidates.extend(_ocr_den4_to_7_expr_candidates(raw))
    s_full7 = _apply_den7_ocr_full_correction(raw)
    if s_full7 != raw:
        candidates.append(s_full7)
    if re.search(r"5\s+1\s*/\s*4", raw):
        s_paren = re.sub(r"\(\s*-?\s*3\s+1\s*/\s*4\s*\)", "(-3 1/7)", raw, count=1)
        if s_paren != raw:
            candidates.append(s_paren)
        s_paren2 = re.sub(r"-\s*\(\s*-?\s*3\s+1\s*/\s*4\s*\)", "-(-3 1/7)", raw, count=1)
        if s_paren2 != raw:
            candidates.append(s_paren2)
    if _student_work_implies_one_and_one_third(work):
        for m in re.finditer(r"\+(\s*)1(\s*/\s*)3(?!\d)", raw):
            s = raw[: m.start()] + "+1 1/3" + raw[m.end() :]
            if s != raw:
                candidates.append(s)
    # 去重保持顺序
    seen = set()
    uniq = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    picked = _pick_best_ocr_expr_candidate(uniq, raw, sv, work)
    if picked:
        return picked
    # 学生过程里已写出正确分母时，允许在多个数值匹配候选中取「过程佐证」的那一项
    for cand in uniq:
        if cand == raw:
            continue
        for m_wrong, m_right in zip(_MIXED_FRAC_TOKEN_RE.finditer(raw), _MIXED_FRAC_TOKEN_RE.finditer(cand)):
            if m_wrong.group(0) != m_right.group(0) and _student_work_has_mixed_frac(
                work, m_right.group(1), m_right.group(2), m_right.group(3)
            ):
                v = safe_eval_math_expr(cand)
                if v is not None and values_equal(v, sv):
                    return cand
    return raw


def convert_mixed_fractions(text: str) -> str:
    """把手写常见的「带分数」如 1 1/3、-32 1/3 转为 (4)/3、(-97)/3，便于后续安全求值。"""
    if not text:
        return ""
    s = str(text).replace("（", "(").replace("）", ")")
    pattern = re.compile(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])")

    def repl(m: re.Match) -> str:
        whole, n, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if d == 0:
            return m.group(0)
        sign = -1 if whole < 0 else 1
        wa = abs(whole)
        num = sign * (wa * d + n)
        return f"({num})/{d}"

    prev = None
    while prev != s:
        prev = s
        s = pattern.sub(repl, s)
    return s


def is_symbolic_expression(expr: str) -> bool:
    """整式化简、解方程等含字母书写：不能做纯数字等号链本地验算，也不能套用横式数值「标准答案」覆盖。"""
    if not expr or not str(expr).strip():
        return False
    raw = str(expr)
    if not re.search(r"[A-Za-z]", raw):
        return False
    compact = re.sub(r"\s+", "", raw.lower())
    if re.fullmatch(r"-?[\d.]+e[+\-]?\d+", compact):
        return False
    return True


def is_radical_expression_context(expr: str, student_work: str = "") -> bool:
    """含平方根、立方根、绝对值等：禁止剥离根号后当纯四则运算求值（否则 √49+∛(-27)… 会变成 19）。"""
    raw = f"{expr or ''}\n{student_work or ''}"
    if not raw.strip():
        return False
    if re.search(
        r"√|∛|∜|sqrt|Sqrt|cbrt|Cbrt|立方根|平方根|绝对值|\|[^|]+\||\*\*\s*0?\s*\.?\s*5",
        raw,
        re.I,
    ):
        return True
    if re.search(r"\\sqrt|\\cbrt|\\dfrac", raw):
        return True
    return False


def uses_non_arithmetic_grading(expr: str, student_work: str = "") -> bool:
    """不走「去根号当整数加减」的本地算术链，含根式/绝对值/代数/方程等。"""
    return (
        is_symbolic_expression(expr)
        or is_algebraic_work_context(expr, student_work)
        or is_radical_expression_context(expr, student_work)
    )


def _sympy_radical_locals():
    from sympy import Abs, Rational, sign, sqrt, sympify

    def real_cbrt(x):
        xv = sympify(x)
        return sign(xv) * (Abs(xv) ** Rational(1, 3))

    return {"sqrt": sqrt, "cbrt": real_cbrt, "Abs": Abs, "Rational": Rational}


def _prepare_radical_expr_text(text: str) -> str:
    t = preprocess_chinese_math_text(str(text or ""))
    t = convert_percent_literals(t)
    t = t.replace("$", "")
    t = re.sub(r"\\sqrt\s*\[\s*3\s*\]\s*\{([^{}]+)\}", r"cbrt(\1)", t)
    t = re.sub(r"\\sqrt\s*\{([^{}]+)\}", r"sqrt(\1)", t)
    t = re.sub(r"\\cbrt\s*\{([^{}]+)\}", r"cbrt(\1)", t)
    t = re.sub(r"\\left|\\right", "", t)
    t = t.replace("×", "*").replace("÷", "/").replace("（", "(").replace("）", ")")
    t = re.sub(r"√\s*\(", "sqrt(", t)
    t = re.sub(r"√\s*(\d+)", r"sqrt(\1)", t)
    t = re.sub(r"∛\s*\(", "cbrt(", t)
    t = re.sub(r"∛\s*(-?\d+)", r"cbrt(\1)", t)
    t = re.sub(r"\|([^|]+)\|", r"Abs(\1)", t)
    t = re.sub(r"\(\s*sqrt\(([^)]+)\)\s*\)\s*\^\s*2", r"(sqrt(\1))**2", t)
    t = re.sub(r"\(\s*sqrt\(([^)]+)\)\s*\)\s*\*\*\s*2", r"(sqrt(\1))**2", t)
    t = re.sub(r"sqrt\(([^)]+)\)\s*\^\s*2", r"(sqrt(\1))**2", t)
    return t


def eval_radical_expression(expr: str):
    """根式/绝对值算式求值；失败返回 None（勿回退到 safe_eval_math_expr）。"""
    if not expr or not is_radical_expression_context(expr):
        return None
    try:
        from sympy import sympify, simplify

        prepared = _prepare_radical_expr_text(expr)
        val = sympify(prepared, locals=_sympy_radical_locals())
        val = simplify(val)
        if val.is_number and val.is_real:
            return Fraction(str(val.evalf())).limit_denominator(10**8)
        return val
    except Exception:
        return None


def format_radical_answer(value) -> str:
    if value is None:
        return ""
    try:
        from sympy import srepr, latex, sympify

        v = sympify(value)
        if v.is_number and v.is_real:
            f = Fraction(str(v.evalf())).limit_denominator(10**8)
            return fraction_to_text(f) if f.denominator != 1 else str(f.numerator)
        s = str(v).replace("sqrt", "√").replace("**", "^").replace("*", "×")
        s = re.sub(r"Abs\(([^)]+)\)", r"|\1|", s)
        return s
    except Exception:
        return str(value)


def extract_radical_answer(text: str):
    """提取根式题最终答案（纯数或含 √ 的式子）。"""
    if not text:
        return None
    raw = str(text).strip()
    if "=" in raw:
        raw = raw.split("=")[-1].strip()
    raw = raw.split("。")[0].split("，")[0].strip()
    if re.search(r"√|sqrt|∛|cbrt|\|", raw, re.I):
        try:
            from sympy import sympify, simplify

            prepared = _prepare_radical_expr_text(raw)
            val = simplify(sympify(prepared, locals=_sympy_radical_locals()))
            return val
        except Exception:
            return re.sub(r"\s+", "", raw)
    return extract_math_value(raw)


def radical_answers_equivalent(a, b) -> bool:
    if a is None or b is None:
        return False
    try:
        if isinstance(a, Fraction) and isinstance(b, Fraction):
            return values_equal(a, b)
        from sympy import simplify, sympify

        da = sympify(a) if not hasattr(a, "is_number") else a
        db = sympify(b) if not hasattr(b, "is_number") else b
        if da.is_number and db.is_number and da.is_real and db.is_real:
            return values_equal(Fraction(str(da.evalf())).limit_denominator(10**8),
                                Fraction(str(db.evalf())).limit_denominator(10**8))
        return simplify(da - db) == 0
    except Exception:
        return re.sub(r"\s+", "", str(a)) == re.sub(r"\s+", "", str(b))


def radical_solution_correct(
    expr: str,
    student_work: str,
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
) -> bool:
    """根式计算题：按 sympy 重算或比对学生最终式，并认可评语已肯定全对的情形。"""
    if not is_radical_expression_context(expr, student_work):
        return False
    standard = eval_radical_expression(expr)
    student = extract_radical_answer(student_answer) or extract_radical_answer(student_work)
    expected = extract_radical_answer(expected_answer)
    if standard is not None and student is not None and radical_answers_equivalent(student, standard):
        return True
    if expected is not None and student is not None and radical_answers_equivalent(student, expected):
        return True
    if reason_claims_student_fully_correct(reason):
        if student is not None:
            if standard is not None and radical_answers_equivalent(student, standard):
                return True
            if expected is not None and radical_answers_equivalent(student, expected):
                return True
        if reason_has_contradictory_false_negative(reason):
            return True
        if str(student_work or "").strip() and re.search(r"[=＝]", str(student_work)):
            return True
        if str(student_answer or "").strip():
            return True
    return False


def sanitize_reason_text(reason: str) -> str:
    """去掉面向学生的订正中不应出现的「本地复核」等元信息，压缩空白。"""
    if not reason:
        return ""
    t = strip_false_negative_grading_hints(str(reason))
    t = re.sub(r"本地复核[^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"复核[:：][^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"说明[:：][^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"与当前作答不一致[^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"按题干重算[^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"结果分记\s*0\s*分[^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"作为[^。！？；;]{0,40}(?:模型|AI)[^。！？；;]*[。！？；;]", "", t)
    t = re.sub(r"(?:OCR|识别器|置信度)[^。！？；;]*[。！？；;]", "", t)
    if "【错误原因】" in t or "【正确步骤】" in t:
        t = re.sub(r"[ \t\f\v]{2,}", " ", t)
        t = re.sub(r"\n{3,}", "\n\n", t).strip()
    else:
        t = re.sub(r"\s+", " ", t).strip()
        t = re.sub(r"[。！？]{2,}", "。", t)
        chunks = [c for c in t.split("正确步骤：") if c.strip()]
        if len(chunks) > 2:
            head = chunks[0].strip()
            tail = chunks[-1].strip()
            t = f"{head} 正确步骤：{tail}".strip()
    return t.strip(" ，。；")


def _equation_or_algebra_hint(expr: str, student_work: str) -> bool:
    """识别解方程、配方、多根等书写：避免被「纯算术等号链」本地校验误伤。"""
    raw = f"{expr or ''}\n{student_work or ''}"
    if not raw.strip():
        return False
    if re.search(r"(?i)(配方|因式分解|开方|方程|未知数|一元|二次|求根|两根|移项|代入|判别式)", raw):
        return True
    if "±" in raw or "∓" in raw:
        return True
    if re.search(r"(?i)\([xy]\s*[\+\-]", raw):
        return True
    if re.search(r"(?i)\bx[_\s]*[₁₂₃₄₅₆₇₈₉₀1234567890]\s*[=＝]", raw):
        return True
    if re.search(r"(?i)\bx\d\s*[=＝]", raw):
        return True
    # 多步书写且出现未知数（排除仅有 × 被误读为 x 的极少数情况：要求有 = 且像方程）
    if raw.count("=") + raw.count("＝") >= 2:
        low = raw.lower()
        if "x" in low or "y" in low:
            return True
    return False


def is_algebraic_work_context(expr: str, student_work: str) -> bool:
    """含未知数、指数、开方±、多根标注等的作答，不能用纯算术「等号链逐段求值」复核，否则会误判。"""
    raw = f"{expr or ''}\n{student_work or ''}"
    if not raw.strip():
        return False
    if _equation_or_algebra_hint(expr, student_work):
        return True
    if re.search(r"[\^⁰¹²³⁴⁵⁶⁷⁸⁹₂₃₄₅₆₇₈₉₀]", raw):
        return True
    if "±" in raw or "∓" in raw:
        return True
    if re.search(r"(?:方程|配方|因式分解|开方|未知数|一元)", raw):
        return True
    # x₁ x_1 x2（下标根）等
    if re.search(r"(?i)\bx[_\s]*[₁₂₃₀1234]|\bx\s*[=＝]", raw):
        return True
    if re.search(r"(?i)(?:\(|\s)[xyXY]\s*[+\-^=)]|[=＝]\s*[xyXY]\s*[^0-9]", raw):
        return True
    return False


def safe_eval_math_expr(expr: str):
    text = str(expr or "")
    text = text.replace("（", "(").replace("）", ")")
    text = preprocess_chinese_math_text(text)
    text = convert_percent_literals(text)
    text = convert_mixed_fractions(text)
    text = rewrite_div_fraction_pairs(text)
    text = normalize_math_expr(text)
    if not text or re.search(r"[=]", text):
        return None
    if not re.fullmatch(r"[0-9+\-*/().\s]+", text):
        return None
    try:
        tree = ast.parse(text, mode="eval")
        return _safe_eval_ast(tree)
    except Exception:
        return None


_NUMERIC_ABS_TOL = 1e-9
_NUMERIC_REL_TOL = 1e-6
_NUMERIC_FRAC_TOL = Fraction(1, 1_000_000_000)


def _decimal_literal_places(value) -> int | None:
    """从十进制字面量推断有效小数位数（用于 13.090909 与 144/11 等等价判定）。"""
    s = str(value).strip()
    if not s or "/" in s:
        return None
    m = re.search(r"\.(\d+)", s)
    return len(m.group(1)) if m else None


def values_equal(a, b, tolerance=None) -> bool:
    """分数/小数/整数等价比较：格式不同但数值相等即视为相同。"""
    if a is None or b is None:
        return False
    if tolerance is None:
        tolerance = _NUMERIC_FRAC_TOL
    try:
        fa, fb = Fraction(a), Fraction(b)
        if fa == fb:
            return True
        if abs(fa - fb) <= tolerance:
            return True
        da, db = float(fa), float(fb)
        if math.isfinite(da) and math.isfinite(db):
            scale = max(abs(da), abs(db), 1.0)
            if abs(da - db) <= max(_NUMERIC_ABS_TOL, _NUMERIC_REL_TOL * scale):
                return True
            for raw in (a, b):
                places = _decimal_literal_places(raw)
                if places is not None and abs(fa - fb) <= Fraction(1, 10**places):
                    return True
        return False
    except Exception:
        return False


def _unique_numeric_values(values):
    """去重：数值等价只保留一个代表。"""
    out = []
    for v in values:
        if v is None:
            continue
        if any(values_equal(v, u) for u in out):
            continue
        out.append(v)
    return out


def collect_comparable_standards(resolved: dict, expected_answer: str = ""):
    """汇总可用于判分的标准值：精确值、四舍五入值、expected_answer 中的数值。"""
    standards = []
    for key in ("exact", "final"):
        v = (resolved or {}).get(key)
        if v is not None:
            standards.append(v)
    exp_raw = str(expected_answer or "").strip()
    if exp_raw:
        ev = extract_math_value(exp_raw)
        if ev is not None:
            standards.append(ev)
        for tok in re.findall(
            r"(?:[+\-])?\d+(?:\.\d+)?\s*/\s*(?:[+\-])?\d+(?:\.\d+)?|(?:[+\-])?\d+(?:\.\d+)?",
            convert_mixed_fractions(preprocess_chinese_math_text(exp_raw)),
        ):
            v = _parse_numeric_token(tok.replace(" ", ""))
            if v is not None:
                standards.append(v)
    return _unique_numeric_values(standards)


def matches_any_standard(candidate, standards) -> bool:
    if candidate is None or not standards:
        return False
    return any(values_equal(candidate, s) for s in standards)


def final_answer_acceptable(
    expr: str,
    student_work: str,
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
) -> bool:
    """最终得数与标准值（分数/小数/整数互通）一致，可豁免末步等号链形式差异。"""
    return bool(
        numeric_outcome_matches(
            expr, student_work, student_answer, expected_answer, reason
        ).get("match")
    )


# 多步分数/小数混合题：合并常数时略有近似，最终与严格通分差在初中阅卷可接受范围内
_MULTISTEP_ADDSUB_SLACK = Fraction(1, 4)


def _expr_allows_numeric_slack(expr: str) -> bool:
    return len(re.findall(r"[+\-]", str(expr or ""))) >= 3


def opposite_sign_final_only(expr: str, student_answer: str, student_work: str) -> bool:
    """最终结果仅为标准值的相反数（典型：除法变乘时丢负号），不按「第一步从原式整体错」清零。"""
    if is_algebraic_work_context(expr, student_work):
        return False
    sv = safe_eval_math_expr(expr)
    av = extract_math_value(student_answer) or extract_math_value(student_work)
    if sv is None or av is None:
        return False
    return av == -sv


_NUMERIC_TOKEN_PATTERN = (
    r"(?:[+\-])?\d+(?:\.\d+)?\s*/\s*(?:[+\-])?\d+(?:\.\d+)?|(?:[+\-])?\d+(?:\.\d+)?"
)


def _parse_numeric_tokens_from_text(normalized: str) -> list:
    if not normalized:
        return []
    return re.findall(_NUMERIC_TOKEN_PATTERN, normalized)


def _pick_best_numeric_token(tokens, expr: str = "", reference_value=None, raw_src: str = ""):
    """取最右侧可解析数值（脱式末步/最终答案）。"""
    if not tokens:
        return None
    for tok in reversed(tokens):
        v = _parse_numeric_token(tok.replace(" ", ""))
        if v is None:
            continue
        if reference_value is not None and expr:
            recovered = recover_decimal_ocr_value(reference_value, v, expr, raw_src)
            if recovered is not None:
                return recovered
        return v
    return None


def extract_final_numeric_value(text: str, expr: str = "", reference_value=None):
    """从脱式/答案文本中提取最终数值（优先等号右侧、分数优于中间整数）。"""
    if not text:
        return None
    raw_src = preprocess_decimal_ocr_text(str(text).replace("（", "(").replace("）", ")"))
    raw = convert_mixed_fractions(preprocess_chinese_math_text(raw_src))
    if "=" in raw:
        tail = raw.split("=")[-1].strip()
        normalized_tail = normalize_math_expr(tail)
        v = _pick_best_numeric_token(
            _parse_numeric_tokens_from_text(normalized_tail), expr, reference_value, raw_src
        )
        if v is not None:
            return v
    normalized = normalize_math_expr(raw)
    if "=" not in normalized:
        ev = safe_eval_math_expr(normalized)
        if ev is not None:
            if reference_value is not None and expr:
                recovered = recover_decimal_ocr_value(reference_value, ev, expr, raw_src)
                if recovered is not None:
                    return recovered
            return ev
    return _pick_best_numeric_token(
        _parse_numeric_tokens_from_text(normalized), expr, reference_value, raw_src
    )


def is_numeric_eval_algebra_task(expr: str, student_work: str, expected_answer: str = "") -> bool:
    """化简后求值、代入等：最终应比对数值，而非代数式字符串。"""
    if not is_algebraic_work_context(expr, student_work):
        return False
    raw = f"{expr or ''}\n{student_work or ''}\n{expected_answer or ''}"
    if re.search(r"求值|化简|代入|已知|当\s*[xyXY]", raw):
        return True
    if re.search(r"[\^²³]|[|｜]", expr or ""):
        return True
    if re.search(r"(?i)[xy]\s*[=＝]", student_work or ""):
        return True
    if extract_math_value(str(expected_answer or "").strip()) is not None and re.search(
        r"(?i)[xy]", expr or ""
    ):
        return True
    return False


def algebra_numeric_outcome_ok(expr: str, student_work: str, student_answer: str, expected_answer: str) -> bool:
    """化简求值类：若最终数值与标准值等价（分数/小数/整数互通），则视为算术上成立。"""
    if not is_numeric_eval_algebra_task(expr, student_work, expected_answer):
        return False
    if looks_like_copied_other_question(expr, student_work, student_answer) or looks_like_only_copied_question(
        expr, student_work
    ):
        return False
    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer)
    standards = collect_comparable_standards(resolved, expected_answer)
    sa = (
        extract_final_numeric_value(student_answer, expr)
        or extract_final_numeric_value(student_work, expr)
        or extract_math_value(student_answer, expr)
        or extract_math_value(student_work, expr)
    )
    if sa is None:
        return False
    if matches_any_standard(sa, standards):
        return True
    if numeric_outcome_matches(expr, student_work, student_answer, expected_answer).get("match"):
        return True
    return False


def _parse_numeric_token(token: str):
    token = str(token or "").replace(" ", "").lstrip("+")
    if not token:
        return None
    try:
        if "/" in token:
            numerator, denominator = token.split("/", 1)
            return Fraction(Decimal(numerator)) / Fraction(Decimal(denominator))
        return Fraction(Decimal(token))
    except (InvalidOperation, ZeroDivisionError, ValueError):
        return None


def _extract_number_tokens(text: str) -> list:
    """提取表达式中的数值 token（不把 1.25 拆成 1 与 25）。"""
    if not text:
        return []
    raw = preprocess_chinese_math_text(str(text).replace("（", "(").replace("）", ")"))
    norm = normalize_math_expr(raw)
    if not norm:
        return []
    return re.findall(r"-?\d+\.\d+|-?\d+", norm)


def _expr_uses_decimal_numbers(expr: str) -> bool:
    return bool(re.search(r"\d+\.\d+", str(expr or "")))


def is_decimal_mul_context(expr: str, student_work: str = "", reason: str = "") -> bool:
    raw = f"{expr or ''}\n{student_work or ''}\n{reason or ''}"
    return _expr_uses_decimal_numbers(expr) or _expr_uses_decimal_numbers(student_work) or bool(
        re.search(r"小\s*数|积的近似|四舍五入|≈|保留.{0,6}位", raw)
    )


def format_answer_display(value, expr: str = "") -> str:
    """小数卷用小数展示；分数/带分数卷保留分数形式，避免把 144/11 显示成 13.090909。"""
    if value is None:
        return ""
    try:
        f = Fraction(value)
    except Exception:
        return str(value)
    if is_decimal_mul_context(expr):
        s = format(float(f), "f").rstrip("0").rstrip(".")
        return s if s else "0"
    return fraction_to_text(f)


_ELEMENTARY_ERROR_TAGS = (
    "【计算结果错误】",
    "【运算顺序错误】",
    "【小数点位置错误】",
    "【取值近似错误】",
    "【解题逻辑错误】",
)


def is_approximate_product_context(expr: str, student_work: str = "", reason: str = "") -> bool:
    raw = f"{expr or ''}\n{student_work or ''}\n{reason or ''}"
    return bool(re.search(r"积的近似|保留.{0,8}位\s*小\s*数|≈|四舍五入", raw))


def format_expr_for_display(expr: str) -> str:
    """展示用：×/÷ 替换时保留带分数里的分数斜杠（禁止把 1 1/3 变成 1 1÷3）。"""
    if not expr:
        return ""
    text = str(expr)
    shields: list[str] = []

    def _shield(m: re.Match) -> str:
        shields.append(m.group(0))
        return f"__MFRAC{len(shields) - 1}__"

    text = re.sub(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])", _shield, text)
    text = text.replace("*", "×").replace("/", "÷")
    for i, orig in enumerate(shields):
        text = text.replace(f"__MFRAC{i}__", orig)
    return text


def ensure_decimal_only_presentation(text: str, expr: str = "", student_work: str = "") -> str:
    """小数卷订正：把 reason/步骤中的分数写法统一换成小数。"""
    if not text or not is_decimal_mul_context(expr, student_work, text):
        return text
    t = format_expr_for_display(str(text))

    def _repl_mixed(m):
        whole, num, den = m.group(1), m.group(2), m.group(3)
        v = _parse_numeric_token(f"{whole} {num}/{den}")
        return format_answer_display(v, expr) if v is not None else m.group(0)

    def _repl_frac(m):
        v = _parse_numeric_token(f"{m.group(1)}/{m.group(2)}")
        return format_answer_display(v, expr) if v is not None else m.group(0)

    t = re.sub(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])", _repl_mixed, t)
    t = re.sub(r"(?<![\d/.])(-?\d+)\s*/\s*(-?\d+)(?![\d/.])", _repl_frac, t)
    return t


def _pick_primary_error_tag(text: str) -> str:
    for tag in _ELEMENTARY_ERROR_TAGS:
        if tag in str(text or ""):
            return tag
    return ""


def classify_arithmetic_error_reason(
    expr: str,
    student_work: str,
    student_answer: str,
    expected_answer: str = "",
    reason: str = "",
) -> str:
    """小数计算题：返回单一主因标签 + 短说明（不含正确步骤）。"""
    outcome = numeric_outcome_matches(expr, student_work, student_answer, expected_answer, reason)
    if outcome.get("match"):
        return ""
    if outcome.get("kind") == "decimal_point":
        return str(outcome.get("message") or "").rstrip("。")
    if detect_expression_structure_violation(expr, student_work):
        return "【运算顺序错误】括号或乘除顺序与题目不一致，应先算括号内再算括号外"
    if first_step_invalid(expr, student_work) and not student_solution_arithmetically_correct(
        expr, student_work, student_answer, expected_answer
    ):
        return "【运算顺序错误】第一步变形或运算顺序与题目不符"
    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer, reason)
    standard = resolved["exact"] or outcome.get("standard")
    final_standard = resolved["final"] or standard
    student = outcome.get("student") or extract_math_value(student_answer) or extract_math_value(student_work)
    places = resolved["places"]
    if places is not None and standard is not None and student is not None:
        rounded = round_to_decimal_places(standard, places)
        if rounded is not None and not values_equal(student, rounded):
            if values_equal(student, standard) and not resolved["is_approximate"]:
                return ""
            fs = format_answer_display(rounded, expr)
            ft = format_answer_display(student, expr)
            exact = format_answer_display(standard, expr)
            return f"【取值近似错误】应先算精确积 {exact}，再保留 {places} 位小数为 {fs}，你写成了 {ft}"
    if final_standard is not None and student is not None and not values_equal(student, final_standard):
        dec = diagnose_decimal_point_error(
            final_standard, student, expr, student_answer or student_work
        )
        if dec:
            return dec.rstrip("。")
        fs = format_answer_display(final_standard, expr)
        ft = format_answer_display(student, expr)
        return f"【计算结果错误】运算结果应为 {fs}，你写成了 {ft}"
    return ""


def consolidate_decimal_correction_reason(
    expr: str,
    student_work: str,
    student_answer: str,
    status: str,
    reason: str,
    expected_answer: str = "",
) -> str:
    """统一小数卷错误评语：单标签 + 正确步骤（全小数）。"""
    if str(status).strip() != "错误":
        return ensure_decimal_only_presentation(
            strip_false_negative_grading_hints(str(reason or "")), expr, student_work
        )
    t = strip_false_negative_grading_hints(str(reason or ""))
    for noisy in (
        r"说明[:：][^。]*[。]",
        r"按题干重算[^。]*[。]",
        r"与当前作答不一致[^。]*[。]",
        r"结果分记\s*0\s*分[^。]*[。]?",
        r"【过程正确、结果错误】[^。]*[。]?",
        r"过程等号在末步不成立[^。]*[。]",
        r"不能判为正确[^。]*[。]",
        r"提示：第一步已从原式错误变形[^。]*[。]",
        r"末步结果与标准答案不一致[^。]*[。]?",
        r"订正提示[：:][^。]*[。]?",
        r"已按规则计为错误[^。]*[。]?",
        r"保留部分过程分[^。]*[。]?",
    ):
        t = re.sub(noisy, "", t)
    if "【小数点位置错误】" in t and "【过程正确、结果错误】" in t:
        t = re.sub(r"【过程正确、结果错误】[^。]*[。]?", "", t)
    classified = classify_arithmetic_error_reason(
        expr, student_work, student_answer, expected_answer, t
    )
    if classified:
        body = classified.rstrip("。")
    else:
        tag = _pick_primary_error_tag(t)
        body = t.strip()
        if tag:
            idx = body.find(tag)
            body = body[idx:].split("正确步骤")[0].split("【正确步骤】")[0].strip()
        elif body:
            body = body.split("正确步骤")[0].split("【正确步骤】")[0].strip()
        else:
            body = "【计算结果错误】运算结果与标准答案不一致"
        for other in _ELEMENTARY_ERROR_TAGS:
            if other != tag and other in body:
                body = body.replace(other, "")
    steps = build_correct_steps(expr, student_work, reason)
    if steps and steps not in body:
        body = f"{body.rstrip('。')}。{steps}"
    return ensure_decimal_only_presentation(body, expr, student_work)


def infer_rounding_decimal_places(expr: str, student_work: str, reason: str = ""):
    raw = f"{expr or ''}\n{student_work or ''}\n{reason or ''}"
    cn = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6}
    m = re.search(r"保留\s*([一二两三1-9])\s*位\s*小\s*数", raw)
    if m:
        ch = m.group(1)
        places = int(ch) if ch.isdigit() else cn.get(ch)
        if places is not None:
            return places
    if re.search(r"两位\s*小\s*数|二位\s*小\s*数|保留两位", raw):
        return 2
    if re.search(r"一位\s*小\s*数|保留一位", raw):
        return 1
    if re.search(r"≈|近似|四舍五入|积的近似", raw):
        steps = split_equation_steps(student_work)
        if len(steps) >= 2:
            for places in (1, 2, 3):
                left_v = safe_eval_math_expr(steps[-2])
                right_v = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
                if left_v is not None and right_v is not None:
                    rounded = round_to_decimal_places(left_v, places)
                    if rounded is not None and values_equal(right_v, rounded):
                        return places
    return None


def _infer_places_from_student_outcome(exact, student_work: str, student_answer: str):
    """从学生最终得数反推保留位数（积的近似数：精确积仅作过程，末位近似值为结果）。"""
    if exact is None:
        return None
    candidates = collect_candidate_answer_values(student_work, student_answer)
    if not candidates:
        return None
    student_final = candidates[-1]
    if values_equal(student_final, exact):
        return None
    matched = []
    for places in (1, 2, 3):
        rounded = round_to_decimal_places(exact, places)
        if rounded is not None and values_equal(student_final, rounded):
            matched.append(places)
    steps = split_equation_steps(student_work)
    if len(steps) >= 2:
        left_v = safe_eval_math_expr(steps[-2])
        right_v = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
        if left_v is not None and values_equal(left_v, exact) and right_v is not None:
            for places in (1, 2, 3):
                rounded = round_to_decimal_places(left_v, places)
                if rounded is not None and values_equal(right_v, rounded):
                    return places
    if len(matched) == 1:
        return matched[0]
    return matched[0] if matched else None


def resolve_grading_standard(
    expr: str,
    student_work: str = "",
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
) -> dict:
    """解析本题评分标准：精确积 + 按要求四舍五入后的最终正确答案。"""
    if is_radical_expression_context(expr, student_work):
        exact = eval_radical_expression(expr)
        if exact is None:
            exact = extract_radical_answer(expected_answer)
        return {
            "exact": exact,
            "places": None,
            "final": exact,
            "is_approximate": False,
        }
    exact = safe_eval_math_expr(expr)
    if exact is None:
        exact = extract_math_value(expected_answer)
    decimal_task = is_decimal_mul_context(expr, student_work, reason) or is_approximate_product_context(
        expr, student_work, reason
    )
    places = infer_rounding_decimal_places(expr, student_work, reason) if decimal_task else None
    if places is None and exact is not None and decimal_task:
        places = _infer_places_from_student_outcome(exact, student_work, student_answer)
    exp_v = extract_math_value(str(expected_answer or "").strip())
    if exact is not None and exp_v is not None:
        if values_equal(exp_v, exact):
            pass
        elif decimal_task:
            for p in (1, 2, 3):
                rounded = round_to_decimal_places(exact, p)
                if rounded is not None and values_equal(exp_v, rounded):
                    places = p
                    break
    final = exact
    if exact is not None and places is not None and decimal_task:
        rounded = round_to_decimal_places(exact, places)
        if rounded is not None:
            final = rounded
    is_approx = bool(
        decimal_task
        and (
            places is not None
            or is_approximate_product_context(expr, student_work, reason)
            or (
                exact is not None
                and final is not None
                and not values_equal(exact, final)
            )
        )
    )
    return {
        "exact": exact,
        "places": places,
        "final": final,
        "is_approximate": is_approx,
    }


def reason_claims_compliant_rounding(reason: str) -> bool:
    """评语是否已认定「先算精确积再按要求取近似」的作答合规。"""
    if not reason:
        return False
    t = str(reason)
    if re.search(r"符合(?:答题)?要求|符合题意|作法正确|做法正确|保留位数正确", t):
        if re.search(r"精确积|四舍五入|保留.{0,8}位|近似|小数", t):
            return True
    return bool(
        re.search(r"计算正确|作答正确|取值正确", t)
        and re.search(r"精确积|四舍五入|保留.{0,8}位|近似", t)
    )


def grading_final_value(
    expr: str,
    student_work: str = "",
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
):
    """本题用于判分的最终数值（积的近似数为四舍五入后的得数）。"""
    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer, reason)
    return resolved["final"] or resolved["exact"]


def student_approximate_rounding_correct(
    expr: str, student_work: str, student_answer: str = "", reason: str = ""
) -> bool:
    """积的近似数：精确积过程 + 末位四舍五入结果均合规。"""
    resolved = resolve_grading_standard(expr, student_work, student_answer, "", reason)
    if resolved["exact"] is None or resolved["final"] is None:
        return False
    if not resolved["is_approximate"] and not is_decimal_mul_context(expr, student_work, reason):
        return False
    candidates = collect_candidate_answer_values(student_work, student_answer)
    if not candidates:
        return False
    if not any(values_equal(v, resolved["final"]) for v in candidates):
        return False
    if reason_claims_compliant_rounding(reason):
        return True
    if not detect_invalid_equal_steps(student_work, expr, student_work, reason):
        return True
    steps = split_equation_steps(student_work)
    if len(steps) >= 2:
        penult = safe_eval_math_expr(steps[-2])
        if penult is not None and not values_equal(penult, resolved["exact"]):
            return False
    return True


def round_to_decimal_places(value, places):
    if value is None or places is None or places < 0:
        return None
    try:
        d = Decimal(str(float(Fraction(value))))
        q = Decimal("0.1") ** int(places)
        return Fraction(d.quantize(q, rounding=ROUND_HALF_UP))
    except Exception:
        return None


def collect_candidate_answer_values(student_work: str, student_answer: str, expr: str = ""):
    """只取学生最终答案（答案栏或脱式末步），不把中间步的精确积当作学生结果。"""
    vals = []

    def add_value(v):
        if v is None:
            return
        if any(values_equal(v, u) for u in vals):
            return
        vals.append(v)

    if str(student_answer or "").strip():
        add_value(extract_final_numeric_value(student_answer, expr) or extract_math_value(student_answer, expr))
    steps = split_equation_steps(student_work)
    if steps:
        add_value(
            safe_eval_math_expr(steps[-1])
            or extract_final_numeric_value(steps[-1], expr)
            or extract_math_value(steps[-1], expr)
        )
    elif str(student_work or "").strip():
        add_value(extract_final_numeric_value(student_work, expr) or extract_math_value(student_work, expr))
    return vals


def has_handwritten_numeric_content(student_work: str, student_answer: str) -> bool:
    combined = f"{student_work or ''}{student_answer or ''}"
    return bool(re.search(r"\d", combined))


def numeric_outcome_matches(
    expr: str,
    student_work: str,
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
):
    """判定学生数值是否可接受：优先核对四舍五入后的最终得数，精确积仅作过程。"""
    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer, reason)
    standard = resolved["exact"]
    if standard is None:
        return {"match": False, "standard": None, "kind": "unknown"}
    final_standard = resolved["final"] or standard
    places = resolved["places"]
    rounded_std = round_to_decimal_places(standard, places) if places is not None else None
    if rounded_std is not None:
        final_standard = rounded_std
    standards = collect_comparable_standards(resolved, expected_answer)
    if not standards:
        standards = [final_standard]
    candidates = collect_candidate_answer_values(student_work, student_answer, expr)
    if not candidates:
        return {"match": False, "standard": standard, "kind": "no_answer", "final_standard": final_standard}
    raw_answer = str(student_answer or "")
    for v in candidates:
        if matches_any_standard(v, standards):
            kind = "rounded" if resolved["is_approximate"] and not values_equal(v, standard) else "exact"
            return {
                "match": True,
                "standard": standard,
                "kind": kind,
                "student": v,
                "places": places,
                "rounded_standard": final_standard,
                "final_standard": final_standard,
            }
        recovered = recover_decimal_ocr_value(final_standard, v, expr, raw_answer or student_work)
        if recovered is not None and matches_any_standard(recovered, standards):
            return {
                "match": True,
                "standard": standard,
                "kind": "ocr_decimal",
                "student": recovered,
                "places": places,
                "rounded_standard": final_standard,
                "final_standard": final_standard,
            }
        if not resolved["is_approximate"] and values_equal(v, standard):
            return {
                "match": True,
                "standard": standard,
                "kind": "exact",
                "student": v,
                "places": places,
                "final_standard": standard,
            }
        recovered_exact = recover_decimal_ocr_value(standard, v, expr, raw_answer or student_work)
        if (
            recovered_exact is not None
            and not resolved["is_approximate"]
            and matches_any_standard(recovered_exact, standards)
        ):
            return {
                "match": True,
                "standard": standard,
                "kind": "ocr_decimal",
                "student": recovered_exact,
                "places": places,
                "final_standard": standard,
            }
    stu = candidates[-1]
    dec_msg = diagnose_decimal_point_error(final_standard, stu, expr, raw_answer or student_work)
    if dec_msg:
        return {
            "match": False,
            "standard": standard,
            "kind": "decimal_point",
            "student": stu,
            "message": dec_msg,
        }
    return {"match": False, "standard": standard, "kind": "wrong", "student": stu}


def student_answer_matches_expr(expr: str, student_work: str, student_answer: str = "") -> bool:
    """学生最终数值与按题干重算一致（含四舍五入保留小数），视为有效作答。"""
    if not expr:
        return False
    return bool(numeric_outcome_matches(expr, student_work, student_answer).get("match"))


def _digit_core_string(value) -> str:
    """去掉符号与小数点，仅保留数字序列，用于小数点错位诊断。"""
    try:
        f = Fraction(value)
        s = format(f, "f")
    except Exception:
        s = str(value)
    return re.sub(r"[^0-9]", "", s).lstrip("0") or "0"


def _power_of_ten_exponent(n: int):
    """若 n 为 10**k（k>=0），返回 k；否则返回 None。"""
    if n <= 0:
        return None
    exp = 0
    x = int(n)
    while x > 1:
        if x % 10 != 0:
            return None
        x //= 10
        exp += 1
    return exp if x == 1 else None


def _raw_shows_explicit_decimal_for_value(raw_text: str, value) -> bool:
    """卷面原文已写出带小数点的该数值，视为学生主动书写的小数位置（非漏识小数点）。"""
    if not raw_text:
        return False
    raw = preprocess_decimal_ocr_text(str(raw_text))
    try:
        target = Fraction(value)
    except Exception:
        return False
    for tok in re.findall(r"-?\d+\.\d+", raw):
        v = _parse_numeric_token(tok)
        if v is not None and values_equal(v, target):
            return True
    return False


def recover_decimal_ocr_value(
    standard_value,
    ocr_value,
    expr: str = "",
    raw_text: str = "",
):
    """OCR 漏识/误识小数点：当识别值为标准答案的 10^n 倍或数字串重复时，还原为可接受作答。"""
    if standard_value is None or ocr_value is None:
        return None
    if values_equal(standard_value, ocr_value):
        return Fraction(ocr_value)
    if expr and not is_decimal_mul_context(expr):
        return None
    try:
        fs = Fraction(standard_value)
        ft = Fraction(ocr_value)
    except Exception:
        return None
    std_digits = re.sub(r"\D", "", format_answer_display(fs, expr)).lstrip("0") or "0"
    raw_digits = re.sub(r"\D", "", str(raw_text or ""))
    if std_digits and raw_digits.startswith(std_digits):
        rest = raw_digits[len(std_digits) :]
        if rest == std_digits or rest == "0" + std_digits:
            return fs
    if _raw_shows_explicit_decimal_for_value(raw_text, ft):
        return None
    shift = _decimal_point_shift_places(fs, ft)
    if shift is not None and shift > 0:
        corrected = ft / (Fraction(10) ** shift)
        if values_equal(corrected, fs):
            if shift >= 2:
                return corrected
            if shift == 1 and _allow_shift1_decimal_ocr_recovery(fs, expr):
                return corrected
    if shift is not None and shift < 0:
        corrected = ft * (Fraction(10) ** (-shift))
        if values_equal(corrected, fs):
            return corrected
    return None


def _allow_shift1_decimal_ocr_recovery(standard: Fraction, expr: str) -> bool:
    """一位倍数：小数积或三位以内整数乘小数口算可容错；四位整数如 1000↔10000 不自动等同。"""
    if standard.denominator != 1:
        return True
    if _expr_uses_decimal_numbers(expr):
        try:
            return abs(int(standard)) < 1000
        except (OverflowError, ValueError):
            return False
    return False


def is_likely_missed_decimal_ocr(
    standard_value, ocr_value, expr: str = "", raw_text: str = ""
) -> bool:
    """识别值与标准值呈 10^n 倍且可还原时，视为 OCR 小数点问题而非学生移位。"""
    return recover_decimal_ocr_value(standard_value, ocr_value, expr, raw_text) is not None


def rewrite_decimal_ocr_numbers_in_text(text: str, expr: str, reference_value) -> str:
    """将文本中与标准值仅差小数点 OCR 的数字串改写为规范小数。"""
    if not text or reference_value is None or not is_decimal_mul_context(expr):
        return text
    ref = Fraction(reference_value)

    def repl(m):
        raw = m.group(0)
        v = _parse_numeric_token(raw)
        if v is None:
            return raw
        recovered = recover_decimal_ocr_value(ref, v, expr, raw)
        if recovered is not None and not values_equal(v, ref):
            return format_answer_display(recovered, expr)
        return raw

    return re.sub(r"-?\d+(?:\.\d+)?", repl, str(text))


def apply_decimal_ocr_correction(
    expr: str,
    student_work: str,
    student_answer: str,
    expected_answer: str = "",
    reason: str = "",
) -> tuple[str, str]:
    """按题干标准值纠正常见「漏识小数点」的 OCR 数值，返回校正后的过程与答案文本。"""
    if not is_decimal_mul_context(expr, student_work, reason):
        return student_work, student_answer
    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer, reason)
    ref = resolved["final"] or resolved["exact"]
    if ref is None:
        return student_work, student_answer
    work = rewrite_decimal_ocr_numbers_in_text(student_work, expr, ref)
    ans = student_answer
    av = extract_math_value(student_answer)
    if av is not None:
        recovered = recover_decimal_ocr_value(ref, av, expr, student_answer)
        if recovered is not None:
            ans = format_answer_display(recovered, expr)
    elif student_answer.strip():
        work_ans = extract_math_value(student_work)
        if work_ans is not None:
            recovered = recover_decimal_ocr_value(ref, work_ans, expr, student_work)
            if recovered is not None:
                ans = format_answer_display(recovered, expr)
    return work, ans


def _decimal_point_shift_places(standard_value, student_value):
    """学生相对标准答案的小数点偏移位数：正=向右多移，负=向左/漏写。"""
    try:
        fs = Fraction(standard_value)
        ft = Fraction(student_value)
    except Exception:
        return None
    if fs == ft or fs == 0 or ft == 0:
        return None
    if (fs < 0) != (ft < 0):
        return None
    ratio = ft / fs
    if ratio.denominator != 1:
        return None
    q = int(ratio.numerator)
    if q > 0:
        exp = _power_of_ten_exponent(q)
        if exp is not None and exp <= 6:
            return exp
    elif q < 0:
        exp = _power_of_ten_exponent(-q)
        if exp is not None and exp <= 6:
            return -exp
    return None


def _decimal_point_shift_hint(shift: int) -> str:
    """按实际偏移位数生成标准话术（仅用于小数点位置错误）。"""
    if shift == 1:
        return "小数点向右多移了一位"
    if shift == 2:
        return "小数点向右多移了两位"
    if shift == 3:
        return "小数点向右多移了三位"
    if shift == -1:
        return "漏写小数点或少写一位小数"
    if shift == -2:
        return "小数点向左多移了两位"
    if shift < -2:
        return "小数位数判断失误，小数点偏左较多"
    if shift > 3:
        return f"小数点向右多移了{shift}位"
    return ""


def diagnose_decimal_point_error(
    standard_value,
    student_value,
    expr: str = "",
    raw_text: str = "",
    *,
    ocr_forgiving: bool = True,
):
    """小数乘法：仅当学生得数为标准答案的精确 10 倍/10^n 倍时判小数点错位。"""
    if standard_value is None or student_value is None:
        return None
    if expr and not is_decimal_mul_context(expr):
        return None
    if values_equal(standard_value, student_value):
        return None
    if ocr_forgiving and is_likely_missed_decimal_ocr(
        standard_value, student_value, expr, raw_text
    ):
        return None
    try:
        fs = Fraction(standard_value)
        ft = Fraction(student_value)
    except Exception:
        return None
    if fs == 0:
        return None
    shift = _decimal_point_shift_places(fs, ft)
    if shift is None or shift == 0:
        return None
    hint = _decimal_point_shift_hint(shift)
    if not hint:
        return None
    fs_txt = format_answer_display(fs, expr)
    ft_txt = format_answer_display(ft, expr)
    return f"【小数点位置错误】{hint}，应得 {fs_txt}，你写成了 {ft_txt}"


def extract_math_value(text: str, expr: str = "", reference_value=None):
    """从答案文本中提取一个可比较的数值，支持整数、小数、分数与带分数。"""
    return extract_final_numeric_value(text, expr, reference_value)


def split_equation_steps(text: str):
    normalized = normalize_math_expr(text)
    if "=" not in normalized:
        return []
    return [p.strip() for p in normalized.split("=") if p.strip()]


def _student_paren_inner_is_eleven_twelfths(student_work: str) -> bool:
    """卷面括号内已通分为 11/12（第 4 题等典型写法）。"""
    if not student_work:
        return False
    return bool(re.search(r"11\s*/\s*12", str(student_work)))


def safe_eval_work_step(step: str, expr: str = "", student_work: str = "") -> Fraction | None:
    """求值脱式单步；若括号内已写出通分结果 11/12，按「整式除括号结果」理解（非把 11/12 当加项）。"""
    v = safe_eval_math_expr(step)
    if v is None or not _student_paren_inner_is_eleven_twelfths(student_work):
        return v
    m = re.search(r"(-?\d+)\s*[*/]\s*\(([^)]+)\)", normalize_math_expr(step))
    if not m:
        return v
    outer = Fraction(int(m.group(1)))
    inner_txt = m.group(2)
    if not re.search(r"11\s*/\s*12", inner_txt):
        return v
    inner_v = Fraction(11, 12)
    reinterpreted = outer / inner_v
    if v is not None and not values_equal(v, reinterpreted):
        return reinterpreted
    return v


def _is_div_to_mul_rewrite(left: str, right: str, left_v, right_v) -> bool:
    """12÷(11/12) 与 12×12/11 等等价变形。"""
    if left_v is not None and right_v is not None and values_equal(left_v, right_v):
        return True
    ln = normalize_math_expr(left)
    rn = normalize_math_expr(right)
    if re.search(r"\*|×", rn) and re.search(r"/|÷", ln):
        lv = safe_eval_math_expr(left)
        rv = safe_eval_math_expr(right)
        if lv is not None and rv is not None and values_equal(lv, rv):
            return True
    return False


def _equal_step_pair_acceptable(
    left: str,
    right: str,
    left_value,
    right_value,
    expr: str,
    student_work: str,
    standard_value=None,
) -> bool:
    if values_equal(left_value, right_value):
        return True
    lv = safe_eval_work_step(left, expr, student_work)
    rv = safe_eval_work_step(right, expr, student_work)
    if lv is not None and rv is not None and values_equal(lv, rv):
        return True
    if standard_value is not None and values_equal(right_value, standard_value):
        if lv is not None and values_equal(lv, right_value):
            return True
    if _is_div_to_mul_rewrite(left, right, left_value, right_value):
        return True
    return False


def detect_invalid_equal_steps(
    text: str,
    expr: str = "",
    student_work_full: str = "",
    reason: str = "",
    student_answer: str = "",
):
    """检查脱式/横式中相邻等号两侧是否数值相等（仅适用于纯算术链；代数方程勿用）。"""
    work = student_work_full or text
    if is_algebraic_work_context(expr, work) or is_radical_expression_context(expr, work):
        return []
    if final_answer_acceptable(expr, work, student_answer, "", reason):
        return []
    steps = split_equation_steps(text)
    resolved = resolve_grading_standard(expr, work, student_answer, "", reason)
    standard_value = resolved["final"] or resolved["exact"]
    places = resolved["places"] or infer_rounding_decimal_places(expr, work, reason)
    invalid = []
    for left, right in zip(steps, steps[1:]):
        left_value = safe_eval_work_step(left, expr, work)
        right_value = safe_eval_work_step(right, expr, work)
        if left_value is None or right_value is None:
            continue
        if _equal_step_pair_acceptable(left, right, left_value, right_value, expr, work, standard_value):
            continue
        if places is not None:
            rounded = round_to_decimal_places(left_value, places)
            if rounded is not None and values_equal(right_value, rounded):
                continue
        if is_decimal_mul_context(expr, work) and diagnose_decimal_point_error(
            left_value, right_value, expr
        ):
            continue
        invalid.append((left, right, left_value, right_value))
    return invalid


def strip_invalidated_work(text: str) -> str:
    """去掉模型可能识别出的涂改、划掉、作废过程，避免参与评分。"""
    if not text:
        return ""
    valid_lines = []
    invalid_markers = (
        "划掉", "划去", "涂掉", "涂改", "擦掉", "作废", "不要", "错写", "废弃",
        "×掉", "叉掉", "删除", "删去", "抹掉", "改为", "原写", "先写", "旧过程",
        "被划", "划线", "划除", "划了", "擦除", "重写前", "修改前", "无效痕迹"
    )
    keep_markers = ("最终", "最后", "改正", "改成", "保留", "未划掉", "有效")
    for raw_line in str(text).splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "改为" in line or "改成" in line:
            parts = re.split(r"改为|改成", line, maxsplit=1)
            line = parts[-1].strip(" ：:，,。") or line
        if any(marker in line for marker in invalid_markers) and not any(marker in line for marker in keep_markers):
            continue
        line = re.sub(r"（?[^，。；;]*?(?:划掉|划去|涂掉|涂改|擦掉|作废|错写|废弃|被划|划除|修改前)[^，。；;]*?）?", "", line)
        line = re.sub(r"\[[^\]]*?(?:划掉|划去|涂改|作废|无效)[^\]]*?\]", "", line).strip()
        if not line:
            continue
        if re.fullmatch(r"[xX×/\\\-—–_~≈\s]+", line):
            continue
        valid_lines.append(line)
    return "\n".join(valid_lines).strip()


def fraction_to_text(value) -> str:
    if value is None:
        return ""
    try:
        value = Fraction(value)
        if value.denominator == 1:
            return str(value.numerator)
        d = value.denominator
        while d > 1 and d % 10 == 0:
            d //= 10
        if d == 1:
            s = format(value, "f").rstrip("0").rstrip(".")
            return s if s else "0"
        return f"{value.numerator}/{value.denominator}"
    except Exception:
        return str(value)


def build_correct_steps(expr: str, student_work: str = "", reason: str = "") -> str:
    """为可计算横式生成简洁的正确步骤和答案（小数卷全程用小数）。"""
    if is_radical_expression_context(expr, student_work):
        val = eval_radical_expression(expr)
        if val is None:
            return "请按运算顺序化简根式与绝对值，逐步计算并写出最终答案。"
        ans = format_radical_answer(val)
        return f"正确步骤：按题意化简根式与绝对值后合并，答案：{ans}。"
    raw = str(expr or "")
    prepared = convert_mixed_fractions(
        preprocess_chinese_math_text(raw.replace("（", "(").replace("）", ")"))
    )
    standard_value = safe_eval_math_expr(prepared)
    if standard_value is None:
        return "请按题意重新列式，逐步计算并写出最终答案。"
    display_expr = format_expr_for_display(expr)
    normalized = normalize_math_expr(prepared)
    ans_txt = format_answer_display(standard_value, expr)
    div_paren = re.match(r"^(-?\d+)\s*/\s*\((.+)\)\s*$", normalized.replace(" ", ""))
    if div_paren and standard_value is not None:
        outer = div_paren.group(1)
        inner_raw = div_paren.group(2)
        inner_value = safe_eval_math_expr(inner_raw)
        if inner_value not in (None, 0):
            inner_txt = fraction_to_text(inner_value)
            inv_txt = fraction_to_text(Fraction(1) / inner_value)
            mul_txt = f"{outer}×{inv_txt}"
            return (
                f"正确步骤：{display_expr}={outer}÷({inner_txt})={mul_txt}={ans_txt}，答案：{ans_txt}。"
            )
    step_lines = []
    paren_expr = normalized
    for inner in re.findall(r"\(([^()]+)\)", normalized):
        inner_value = safe_eval_math_expr(inner)
        if inner_value is not None:
            paren_expr = paren_expr.replace(
                f"({inner})", format_answer_display(inner_value, expr), 1
            )
    if paren_expr != normalized:
        step_lines.append(format_expr_for_display(paren_expr))
    places = infer_rounding_decimal_places(expr, student_work, reason)
    if places is not None and is_decimal_mul_context(expr, student_work):
        exact_txt = format_answer_display(standard_value, expr)
        rounded = round_to_decimal_places(standard_value, places)
        round_txt = format_answer_display(rounded, expr) if rounded is not None else exact_txt
        if values_equal(standard_value, rounded):
            chain = f"{display_expr}={round_txt}"
        else:
            chain = f"{display_expr}={exact_txt}≈{round_txt}"
        return f"正确步骤：{chain}，答案：{round_txt}。"
    if not step_lines:
        step_lines.append(ans_txt)
    else:
        step_lines.append(ans_txt)
    return f"正确步骤：{display_expr}=" + "=".join(step_lines) + f"，答案：{ans_txt}。"


def first_step_invalid(expr: str, student_work: str, student_answer: str = "", expected_answer: str = "") -> bool:
    """「第一步」错误：第一个等号两侧数值不成立，或第一步左侧整体值已与题干原式不一致。

    不要求第一个等号右侧等于整题最终答案（允许先化简括号内再除等合法中间式）。
    """
    if is_algebraic_work_context(expr, student_work):
        return False
    if final_answer_acceptable(expr, student_work, student_answer, expected_answer):
        return False
    expr_value = safe_eval_math_expr(expr)
    steps = split_equation_steps(student_work)
    if expr_value is None or len(steps) < 2:
        return False
    # 等号链每步成立且末步等于整题标准值（含四舍五入后的近似得数）：允许首行跳步
    if not detect_invalid_equal_steps(student_work, expr, student_work):
        final_val = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
        target = grading_final_value(expr, student_work) or expr_value
        if final_val is not None and target is not None and values_equal(final_val, target):
            return False
        if student_approximate_rounding_correct(expr, student_work, ""):
            return False
    exn = normalize_math_expr(re.sub(r"\s+", "", str(expr)))
    for i in range(len(steps) - 1):
        left_value = safe_eval_math_expr(steps[i])
        right_value = safe_eval_math_expr(steps[i + 1])
        if left_value is None or right_value is None:
            continue
        if not values_equal(left_value, right_value):
            if i != 0:
                return False
            s0 = normalize_math_expr(re.sub(r"\s+", "", steps[i]))
            if s0 == exn:
                return True
            # 左侧数值已达整题标准值，多为中间步骤正确、仅后续抄写/末步算错，不按「第一步从原式错」论处
            if values_equal(left_value, expr_value):
                return False
            return True
    first_left_value = safe_eval_math_expr(steps[0])
    if first_left_value is not None and not values_equal(first_left_value, expr_value):
        # 与 student_solution 一致：多步加减允许微小合并/取整差，不因跳步合并与冗长通分式略有出入判「第一步错」
        if _expr_allows_numeric_slack(expr) and values_equal(first_left_value, expr_value, tolerance=_MULTISTEP_ADDSUB_SLACK):
            return False
        return True
    return False


def looks_like_only_copied_question(expr: str, student_work: str) -> bool:
    """作答仅重复题干/原式，无有效变形或计算（与「未写」区分：有字但未算）。"""
    if not expr or not student_work:
        return False
    if is_algebraic_work_context(expr, student_work):
        return False
    raw = strip_invalidated_work(str(student_work)).strip()
    if not raw:
        return False
    ex = normalize_math_expr(re.sub(r"\s+", "", str(expr)))
    wx = normalize_math_expr(re.sub(r"\s+", "", raw.replace("\n", "")))
    if not ex or not wx:
        return False
    if "=" not in raw:
        return wx == ex or (ex in wx and len(wx) - len(ex) <= 2)
    parts = split_equation_steps(raw)
    if len(parts) < 2:
        seg = normalize_math_expr(parts[0].replace(" ", "")) if parts else wx
        return bool(seg) and seg == ex
    p0 = normalize_math_expr(parts[0].replace(" ", ""))
    p1 = normalize_math_expr(parts[1].replace(" ", ""))
    if p0 == ex and p1 == ex:
        return True
    if p0 == ex:
        v1 = safe_eval_math_expr(parts[1])
        ve = safe_eval_math_expr(expr)
        if v1 is not None and ve is not None and values_equal(v1, ve) and parts[0].strip() == parts[1].strip():
            return True
    return False


def looks_like_copied_other_question(expr: str, student_work: str, student_answer: str = "") -> bool:
    """识别相邻题串题；卷面有手写数字或结果数值可接受时，一律不按串题/未作答处理。"""
    if not expr or not student_work:
        return False
    if has_handwritten_numeric_content(student_work, student_answer):
        return False
    if student_answer_matches_expr(expr, student_work, student_answer):
        return False
    expr_norm = normalize_math_expr(expr)
    work_norm = normalize_math_expr(student_work)
    steps = split_equation_steps(work_norm)
    first = steps[0] if steps else work_norm
    expr_tokens = set(_extract_number_tokens(expr))
    work_tokens = set(_extract_number_tokens(first))
    if expr_tokens and work_tokens and (expr_tokens & work_tokens):
        return False
    if "=" in str(student_work) and expr_norm:
        ex_compact = re.sub(r"\s+", "", expr_norm)
        w_compact = re.sub(r"\s+", "", work_norm)
        if ex_compact in w_compact or w_compact.startswith(ex_compact):
            return False
    expr_ops = set(re.findall(r"[+\-*/]", expr_norm))
    first_ops = set(re.findall(r"[+\-*/]", first))
    if len(expr_tokens) >= 2 and work_tokens and expr_tokens.isdisjoint(work_tokens):
        if _expr_uses_decimal_numbers(expr) or _expr_uses_decimal_numbers(student_work):
            return False
        return True
    if (
        expr_ops
        and first_ops
        and expr_ops.isdisjoint(first_ops)
        and len(expr_tokens & work_tokens) <= 1
        and not _expr_uses_decimal_numbers(expr)
    ):
        return True
    return False


_BLANK_LIKE_TOKENS = frozenset(
    {"", "无", "空", "空白", "未写", "未作答", "没写", "无作答", "没有", "—", "-", "略", "无内容"}
)


def lacks_effective_student_response(expr: str, student_work: str, student_answer: str) -> bool:
    """作答完整性：仅无任何手写数字时视为未作答；识别到数字（含漏小数点）一律视为已作答。"""
    work = strip_invalidated_work(str(student_work or "")).strip()
    ans = strip_invalidated_work(str(student_answer or "")).strip()
    if has_handwritten_numeric_content(work, ans):
        return False
    if work in _BLANK_LIKE_TOKENS and ans in _BLANK_LIKE_TOKENS:
        return True
    return not re.search(r"\d", f"{work}\n{ans}")


def _split_top_level_binop(s: str):
    s = re.sub(r"\s+", "", str(s or ""))
    if not s:
        return None
    depth = 0
    for i in range(len(s) - 1, -1, -1):
        ch = s[i]
        if ch == ")":
            depth += 1
        elif ch == "(":
            depth -= 1
        elif depth == 0 and ch in "*/":
            return s[:i], ch, s[i + 1 :]
    return None


def _parse_simple_side(side: str):
    side = re.sub(r"\s+", "", str(side or ""))
    m = re.fullmatch(r"(\d+)", side)
    if m:
        return {"kind": "num", "value": int(m.group(1))}
    m = re.fullmatch(r"\((\d+)([\+\-])(\d+)\)", side)
    if m:
        return {"kind": "paren_bin", "op": m.group(2), "a": int(m.group(1)), "b": int(m.group(3))}
    return {"kind": "other"}


def _expression_structure_signature(expr: str):
    if not expr or is_algebraic_work_context(expr, ""):
        return None
    prepared = convert_mixed_fractions(
        preprocess_chinese_math_text(str(expr).replace("（", "(").replace("）", ")"))
    )
    prepared = normalize_math_expr(prepared)
    split = _split_top_level_binop(prepared)
    if not split:
        return None
    left, op, right = split
    return {"outer_op": op, "left": _parse_simple_side(left), "right": _parse_simple_side(right)}


def _is_swapped_mul_paren_structure(expected: dict, actual: dict) -> bool:
    if expected.get("outer_op") != "*" or actual.get("outer_op") != "*":
        return False
    el, er = expected["left"], expected["right"]
    al, ar = actual["left"], actual["right"]
    if el.get("kind") != "num" or er.get("kind") != "paren_bin" or er.get("op") not in ("+", "-"):
        return False
    if al.get("kind") != "paren_bin" or al.get("op") not in ("+", "-") or ar.get("kind") != "num":
        return False
    a = el["value"]
    if a not in (al["a"], al["b"]):
        return False
    if ar["value"] not in (er["a"], er["b"]):
        return False
    return True


def _is_swapped_div_paren_structure(expected: dict, actual: dict) -> bool:
    if expected.get("outer_op") != "/" or actual.get("outer_op") != "/":
        return False
    el, er = expected["left"], expected["right"]
    al, ar = actual["left"], actual["right"]
    if el.get("kind") != "num" or er.get("kind") != "paren_bin" or er.get("op") != "-":
        return False
    if al.get("kind") != "paren_bin" or al.get("op") != "+" or ar.get("kind") != "num":
        return False
    return el["value"] in (al["a"], al["b"])


def detect_expression_structure_violation(expr: str, student_work: str) -> bool:
    """学生首步是否把括号/运算顺序改成与原题不同的结构（如 22×(16+21) 写成 (22+16)×21）。"""
    if not expr or not student_work:
        return False
    if is_algebraic_work_context(expr, student_work):
        return False
    ex_sig = _expression_structure_signature(expr)
    if not ex_sig:
        return False
    steps = split_equation_steps(student_work)
    bodies = steps if steps else [student_work]
    for body in bodies[:2]:
        body = str(body or "").strip()
        if not body or len(body) < 4:
            continue
        st_sig = _expression_structure_signature(body)
        if not st_sig:
            continue
        if _is_swapped_mul_paren_structure(ex_sig, st_sig) or _is_swapped_div_paren_structure(ex_sig, st_sig):
            return True
    return False


def is_last_step_only_arithmetic_error(expr: str, student_work: str, invalid_steps) -> bool:
    """仅末步四则算错，前几步（含括号内）结构/数值链仍成立。"""
    if not invalid_steps or len(invalid_steps) != 1:
        return False
    if detect_expression_structure_violation(expr, student_work) or first_step_invalid(expr, student_work):
        return False
    _, _, lv, rv = invalid_steps[0]
    if diagnose_decimal_point_error(lv, rv, expr):
        return False
    places = infer_rounding_decimal_places(expr, student_work, "")
    if places is not None and lv is not None and rv is not None:
        rounded = round_to_decimal_places(lv, places)
        if rounded is not None and values_equal(rv, rounded):
            return False
    steps = split_equation_steps(student_work)
    if len(steps) < 2:
        return False
    left, right, _, _ = invalid_steps[0]
    if right != steps[-1]:
        return False
    if left != steps[-2] and len(steps) >= 2:
        return False
    return True


def apply_process_correct_result_wrong(
    expr: str,
    student_work: str,
    student_answer: str,
    status: str,
    process_score: float,
    result_score: float,
    reason: str,
) -> tuple:
    """过程对、结果错：扣结果分；小数题四舍五入正确或小数点错位时不误用本标签。"""
    if is_algebraic_work_context(expr, student_work):
        return status, process_score, result_score, reason
    if is_radical_expression_context(expr, student_work):
        if student_solution_arithmetically_correct(
            expr, student_work, student_answer, "", reason
        ):
            return (
                "正确",
                max(process_score, 6),
                max(result_score, 3),
                strip_false_negative_grading_hints(reason)
                or "过程、每步等式和最终结果均正确。",
            )
        return status, process_score, result_score, strip_false_negative_grading_hints(reason)
    outcome = numeric_outcome_matches(expr, student_work, student_answer, reason=reason)
    if outcome.get("match"):
        return "正确", max(process_score, 6), max(result_score, 3), strip_false_negative_grading_hints(reason) or (
            "过程、每步等式和最终结果均正确。"
        )
    if student_approximate_rounding_correct(expr, student_work, student_answer, reason):
        return "正确", max(process_score, 6), max(result_score, 3), strip_false_negative_grading_hints(reason) or (
            "先算精确乘积，再按题意四舍五入保留小数位数，作答正确。"
        )
    if outcome.get("kind") == "decimal_point":
        return (
            "错误",
            0,
            0,
            outcome.get("message") or reason,
        )
    resolved = resolve_grading_standard(expr, student_work, student_answer, "", reason)
    standard_value = outcome.get("standard") or resolved["exact"] or safe_eval_math_expr(expr)
    final_standard = resolved["final"] or standard_value
    if standard_value is None:
        return status, process_score, result_score, reason
    answer_value = outcome.get("student") or extract_math_value(student_answer) or extract_math_value(
        student_work
    )
    if answer_value is None or values_equal(answer_value, final_standard) or values_equal(
        answer_value, standard_value
    ):
        return status, process_score, result_score, reason
    dec_hint = diagnose_decimal_point_error(final_standard, answer_value, expr)
    if dec_hint:
        return "错误", 0, 0, dec_hint
    invalid_steps = detect_invalid_equal_steps(student_work, expr, student_work)
    if not invalid_steps and student_solution_arithmetically_correct(expr, student_work, student_answer):
        return (
            "正确",
            max(process_score, 6),
            max(result_score, 3),
            strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。",
        )
    if not is_last_step_only_arithmetic_error(expr, student_work, invalid_steps):
        if not invalid_steps:
            return status, process_score, result_score, reason
        if first_step_invalid(expr, student_work):
            return status, process_score, result_score, reason
    status = "错误"
    result_score = 0
    process_score = max(min(process_score, 6), 4)
    if is_decimal_mul_context(expr, student_work, reason):
        classified = classify_arithmetic_error_reason(
            expr, student_work, student_answer, reason=reason
        )
        if classified:
            reason = f"{classified.rstrip('。')}。{build_correct_steps(expr, student_work, reason)}"
        return status, process_score, result_score, reason
    tag = "【过程正确、结果错误】"
    if tag not in reason and not is_decimal_mul_context(expr, student_work, reason):
        reason = (
            f"{tag} 括号内或前几步变形正确，但最终结果与标准答案 "
            f"{format_answer_display(final_standard, expr)} 不一致。{reason}"
        ).strip()
    return status, process_score, result_score, reason


def remove_invalidated_reason_fragments(reason: str) -> str:
    """从评语中移除把涂改/划掉内容当错误讲解的片段。"""
    if not reason:
        return ""
    text = str(reason)
    text = re.sub(r"[^。！？；;]*?(?:划掉|划去|涂掉|涂改|擦掉|作废|错写|废弃|被划掉|被划去|修改前|旧过程)[^。！？；;]*?[。！？；;]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def reason_has_contradictory_false_negative(reason: str) -> bool:
    """评语前半肯定结果正确，后半又写末步与标准答案不一致等自相矛盾套话。"""
    if not reason:
        return False
    t = re.sub(r"\$+[^$]*\$+", "", str(reason))
    if "结果正确" not in t and "最终结果正确" not in t:
        return False
    return bool(
        re.search(r"订正提示|末步[^。]{0,24}标准答案[^。]{0,12}不一致|已按规则计为错误", t)
    )


def strip_false_negative_grading_hints(reason: str) -> str:
    """去掉与本地验算结论矛盾的「误错」套话（如跳步却判第一步整题错、结果实际一致却称末步不一致）。"""
    if not reason:
        return ""
    t = str(reason)
    t = re.sub(r"\s*提示：第一步已从原式错误变形[^。]*。", "", t)
    t = re.sub(r"[（(]?\s*订正提示[：:][^）)]*?[）)]", "", t)
    t = re.sub(r"订正提示[：:][^。\n]*?末步[^。\n]*?标准答案[^。\n]*[。）)]?", "", t)
    t = re.sub(r"末步结果与标准答案不一致[^。]*[。]?", "", t)
    t = re.sub(r"订正提示[：:][^。]*[。]?", "", t)
    t = re.sub(r"已按规则计为错误[^。]*[。]?", "", t)
    t = re.sub(r"说明：题目数值结果应为[^。]*?与当前作答不一致。", "", t)
    t = re.sub(r"说明：按题干重算标准结果约为[^。]*?与作答过程或最终结果不一致。", "", t)
    t = re.sub(r"得\s*4\s*[又帯带]?\s*1\s*/\s*12[^。，,；;]*[。，,；;]?", "", t)
    t = re.sub(r"4\s*[又帯带]?\s*1\s*/\s*12", "11/12", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def reason_claims_student_fully_correct(reason: str) -> bool:
    """评语是否明确肯定过程/结果/等号链均正确（与「判错」矛盾）。"""
    if not reason:
        return False
    t = re.sub(r"\$+[^$]*\$+", "", str(reason))
    t = t.replace("\\sqrt{2}", "√2").replace("\\sqrt2", "√2")
    if "过程合理且结果正确" in t or "过程合理，结果正确" in t:
        return True
    markers = (
        "结果正确", "过程正确", "过程合理", "过程清晰正确", "等号链成立",
        "每步成立", "最终结果正确", "整体解答结果正确", "运算能力",
        "合并后得", "合并后保留", "合并同类项", "计算正确", "正确计算",
    )
    return sum(1 for m in markers if m in t) >= 2 or (
        "结果正确" in t and ("过程" in t or "等号" in t or "每步" in t)
    ) or ("每步成立" in t and "结果正确" in t) or (
        "结果正确" in t and re.search(r"(?:平方根|立方根|绝对值|根式|√|sqrt)", t, re.I)
    )


def reason_mentions_fabricated_steps(student_work: str, reason: str) -> bool:
    """reason 是否出现 student_work 中找不到的分数/带分数演算（疑似编造过程）。"""
    if not reason or not student_work:
        return False
    # 「正确步骤」为系统/模型模板，其中的分数不必出现在 student_work
    reason_check = str(reason).split("正确步骤")[0].split("【正确步骤】")[0]
    if not reason_check.strip():
        return False
    work = preprocess_chinese_math_text(str(student_work))
    for m in re.finditer(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])", reason_check):
        token = f"{m.group(1)} {m.group(2)}/{m.group(3)}"
        if token not in work and token.replace(" ", "") not in work.replace(" ", ""):
            return True
    for m in re.finditer(r"(?<![\d/.])(-?\d+)\s*/\s*(\d+)(?![\d/.])", reason_check):
        token = f"{m.group(1)}/{m.group(2)}"
        if token not in work and work.count(token) == 0:
            if re.search(rf"{re.escape(m.group(1))}\s*/\s*{re.escape(m.group(2))}", work):
                continue
            return True
    return False


def patch_contradictory_model_verdict(q: dict) -> dict | None:
    """模型评语同时写「结果正确」与「末步与标准答案不一致」时，直接改判正确（短路）。"""
    if not q:
        return None
    raw_reason = str(q.get("reason") or "")
    if not reason_has_contradictory_false_negative(raw_reason):
        return None
    expr = str(q.get("expr", "")).strip()
    work = strip_invalidated_work(str(q.get("student_work", q.get("work", "")) or "").strip())
    ans = strip_invalidated_work(str(q.get("student_answer", q.get("answer", "")) or "").strip())
    if not (reason_claims_student_fully_correct(raw_reason) or is_radical_expression_context(expr, work)):
        return None
    out = dict(q)
    out["expr"] = expr
    out["student_work"] = work
    out["student_answer"] = ans
    out["status"] = "正确"
    out["process_score"] = max(clamp_score(out.get("process_score"), 0, 6), 6)
    out["result_score"] = max(clamp_score(out.get("result_score"), 0, 3), 3)
    out["structure_score"] = max(clamp_score(out.get("structure_score"), 0, 1), 1)
    out["total_score"] = clamp_score(
        out["process_score"] + out["result_score"] + out["structure_score"], 0, 10
    )
    if is_radical_expression_context(expr, work):
        rad_std = eval_radical_expression(expr)
        if rad_std is not None:
            out["expected_answer"] = format_radical_answer(rad_std)
    cleaned = strip_false_negative_grading_hints(raw_reason)
    out["reason"] = sanitize_reason_text(cleaned or "过程、每步等式和最终结果均正确。")
    return out


def enforce_verdict_reason_consistency(
    expr: str,
    student_work: str,
    student_answer: str,
    status: str,
    process_score: float,
    result_score: float,
    structure_score: float,
    reason: str,
    expected_answer: str = "",
) -> tuple[str, float, float, float, str]:
    """末轮兜底：评语已肯定全对或本地验算成立时，不得保留「错误」+ 订正矛盾套话。"""
    raw_reason = str(reason or "").strip()
    contradictory = reason_has_contradictory_false_negative(raw_reason)
    reason = strip_false_negative_grading_hints(raw_reason)
    st = str(status or "").strip()
    arith_ok = student_solution_arithmetically_correct(
        expr, student_work, student_answer, expected_answer, raw_reason
    )
    claims_ok = reason_claims_student_fully_correct(raw_reason)
    radical_ctx = is_radical_expression_context(expr, student_work)
    if contradictory and st in ("错误", "过程不规范"):
        st = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = reason or "过程、每步等式和最终结果均正确。"
    elif (
        st in ("错误", "过程不规范")
        and (arith_ok or claims_ok)
        and (radical_ctx or arith_ok or claims_ok)
        and not looks_like_only_copied_question(expr, student_work)
        and not detect_expression_structure_violation(expr, student_work)
    ):
        st = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = reason or "过程、每步等式和最终结果均正确。"
    return st, process_score, result_score, structure_score, reason


def reconcile_reason_with_verdict(
    reason: str, status: str, arith_ok: bool, student_work: str, expr: str = ""
) -> tuple[str, str]:
    """消除「评语说全对却判错」或 reason 编造过程与卷面不符时的展示矛盾。"""
    reason = strip_false_negative_grading_hints(str(reason or "").strip())
    st = str(status or "").strip()
    if reason_has_contradictory_false_negative(reason) and is_radical_expression_context(expr, student_work):
        arith_ok = True
    if is_radical_expression_context(expr, student_work) and reason_claims_student_fully_correct(reason):
        arith_ok = True
    if arith_ok and st in ("错误", "过程不规范"):
        return "正确", reason_claims_student_fully_correct(reason) and reason or "过程、每步等式和最终结果均正确。"
    if st == "错误" and reason_claims_student_fully_correct(reason) and not arith_ok:
        reason = re.sub(
            r"(过程|结果).{0,8}正确[^。]*[。]",
            "",
            reason,
        ).strip()
    if (
        st == "错误"
        and reason_mentions_fabricated_steps(student_work, reason)
        and not arith_ok
        and not reason_claims_student_fully_correct(reason)
        and "正确步骤" not in str(reason)
    ):
        reason = (
            "卷面该题作答与评语中描述的过程不一致，已按卷面可见书写重新核验。"
            " 若你确认卷面作答完整，可通过「判题有误反馈」提交复核。"
        )
    return st, reason


def student_solution_arithmetically_correct(
    expr: str,
    student_work: str,
    student_answer: str = "",
    expected_answer: str = "",
    reason: str = "",
) -> bool:
    """本地确认横式/脱式是否每步成立且最终值正确。"""
    if looks_like_only_copied_question(expr, student_work):
        return False

    if is_numeric_eval_algebra_task(expr, student_work, expected_answer):
        return algebra_numeric_outcome_ok(expr, student_work, student_answer, expected_answer)

    if is_radical_expression_context(expr, student_work):
        return radical_solution_correct(
            expr, student_work, student_answer, expected_answer, reason
        )

    if is_symbolic_expression(expr):
        return False

    outcome = numeric_outcome_matches(expr, student_work, student_answer, expected_answer)
    if outcome.get("match"):
        return True
    if outcome.get("kind") == "decimal_point":
        return False
    if student_approximate_rounding_correct(expr, student_work, student_answer):
        return True

    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer)
    standard_value = resolved["exact"] or outcome.get("standard") or safe_eval_math_expr(expr)
    final_standard = resolved["final"] or standard_value
    if standard_value is None:
        return False
    invalid_steps = detect_invalid_equal_steps(student_work, expr, student_work)
    if invalid_steps:
        return False
    steps = split_equation_steps(student_work)
    if steps:
        first_value = safe_eval_math_expr(steps[0])
        final_value = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
        if final_value is not None:
            if values_equal(final_value, final_standard):
                return True
            if not resolved["is_approximate"] and values_equal(final_value, standard_value):
                return True
            if _expr_allows_numeric_slack(expr) and values_equal(
                final_value, final_standard, tolerance=_MULTISTEP_ADDSUB_SLACK
            ):
                return True
        if first_value is not None and not values_equal(first_value, standard_value):
            if not (
                _expr_allows_numeric_slack(expr)
                and values_equal(first_value, standard_value, tolerance=_MULTISTEP_ADDSUB_SLACK)
            ):
                return False
        if final_value is not None:
            return False
    answer_value = extract_math_value(student_answer) or extract_math_value(student_work)
    if answer_value is None:
        return False
    if values_equal(answer_value, final_standard):
        return True
    if not resolved["is_approximate"] and values_equal(answer_value, standard_value):
        return True
    if _expr_allows_numeric_slack(expr) and values_equal(
        answer_value, final_standard, tolerance=_MULTISTEP_ADDSUB_SLACK
    ):
        return True
    return False


def local_arithmetic_guard(q: dict, status: str, process_score: float, result_score: float, structure_score: float, reason: str):
    """对 AI 结果做本地算术兜底，防止明显算式误判为正确。"""
    expr = q.get("expr", "")
    student_work = q.get("student_work", "")
    student_answer = q.get("student_answer", "")
    expected_answer = q.get("expected_answer", "")

    resolved = resolve_grading_standard(expr, student_work, student_answer, expected_answer, reason)
    expr_value = resolved["exact"] or safe_eval_math_expr(expr)
    expected_value = extract_math_value(expected_answer)
    standard_value = expr_value if expr_value is not None else expected_value
    final_standard = resolved["final"] or standard_value
    answer_value = extract_math_value(student_answer) or extract_math_value(student_work)
    arith_chain_ok = student_solution_arithmetically_correct(
        expr, student_work, student_answer, q.get("expected_answer", "") or expected_answer
    )
    invalid_steps = detect_invalid_equal_steps(student_work, expr, student_work, reason)
    copied_other_question = looks_like_copied_other_question(expr, student_work, student_answer)
    first_step_wrong = first_step_invalid(expr, student_work)
    correct_steps = build_correct_steps(expr, student_work, reason)
    symbolic = uses_non_arithmetic_grading(expr, student_work)

    if copied_other_question and not has_handwritten_numeric_content(student_work, student_answer):
        status = "未作答"
        process_score = 0
        result_score = 0
        structure_score = 0
        q["student_work"] = ""
        q["student_answer"] = ""
        reason = (f"本道题作答区域疑似为空，识别到的过程与题目数字/运算符不匹配，可能属于相邻题，"
                  f"不作为本题有效作答。{correct_steps}").strip()
        return status, process_score, result_score, structure_score, reason

    if looks_like_only_copied_question(expr, student_work):
        status = "错误"
        process_score = 0
        result_score = 0
        structure_score = 0
        reason = (
            f"【未有效计算】仅重复题干，未见「=」后的计算或答案。{correct_steps}"
        ).strip()
        return status, process_score, result_score, structure_score, reason

    # 模型若已判「正确」但数值/等号链与标准不符，不得在此提前返回，否则末步算错也会被放过。

    if not symbolic and final_standard is not None:
        q["expected_answer"] = format_answer_display(final_standard, expr)
        standards = collect_comparable_standards(resolved, expected_answer)
        if answer_value is not None and not matches_any_standard(answer_value, standards or [final_standard]):
            if not arith_chain_ok:
                status = "错误"
                result_score = 0
                process_score = min(process_score, 0 if first_step_wrong else 5)
                dec_hint = diagnose_decimal_point_error(final_standard, answer_value, expr)
                if dec_hint:
                    reason = f"{dec_hint}。{correct_steps}".strip()
                elif is_decimal_mul_context(expr, student_work, reason):
                    classified = classify_arithmetic_error_reason(
                        expr, student_work, student_answer, expected_answer, reason
                    )
                    reason = (
                        f"{classified or '【计算结果错误】运算结果与标准答案不一致'}。{correct_steps}"
                    ).strip()
                else:
                    exp_ans = format_answer_display(standard_value, expr)
                    reason = (
                        f"{reason} 说明：题目数值结果应为 {exp_ans}，"
                        f"与当前作答不一致。{correct_steps}"
                    ).strip()

    if not symbolic and invalid_steps and not final_answer_acceptable(
        expr, student_work, student_answer, expected_answer, reason
    ):
        left, right, left_value, right_value = invalid_steps[0]
        status = "错误"
        result_score = 0
        dec_pair = diagnose_decimal_point_error(left_value, right_value, expr) or diagnose_decimal_point_error(
            final_standard, answer_value, expr
        )
        if dec_pair:
            process_score = 0
            reason = f"{dec_pair}。{correct_steps}".strip()
        elif is_decimal_mul_context(expr, student_work, reason):
            process_score = 0 if first_step_wrong else max(min(process_score, 6), 4)
            classified = classify_arithmetic_error_reason(
                expr, student_work, student_answer, expected_answer, reason
            )
            reason = f"{classified or dec_pair or '【计算结果错误】末步计算有误'}。{correct_steps}".strip()
        elif is_last_step_only_arithmetic_error(expr, student_work, invalid_steps):
            process_score = max(min(process_score, 6), 4)
            reason = (
                f"【过程正确、结果错误】过程等号在末步不成立：{left}={format_answer_display(left_value, expr)}，"
                f"但 {right}={format_answer_display(right_value, expr)}；标准答案为 "
                f"{format_answer_display(standard_value, expr)}。 {reason}"
            ).strip()
        else:
            process_score = 0 if first_step_wrong else min(process_score, 3)
            reason = (
                f"{reason} 说明：过程等号不成立，{left}={format_answer_display(left_value, expr)}，"
                f"但 {right}={format_answer_display(right_value, expr)}，不能判为正确。{correct_steps}"
            ).strip()
        if (
            final_standard is not None
            and answer_value is not None
            and values_equal(answer_value, final_standard)
        ):
            result_score = min(result_score, 2)

    if status == "错误" and not reason:
        reason = f"作答存在错误。{correct_steps}"
    elif status == "错误" and "正确步骤" not in reason:
        reason = f"{reason} {correct_steps}".strip()

    if status == "正确" and not symbolic:
        outcome = numeric_outcome_matches(expr, student_work, student_answer, expected_answer, reason)
        if outcome.get("match"):
            pass
        elif invalid_steps:
            status = "错误"
            result_score = 0
            left, right, left_value, right_value = invalid_steps[0]
            dec_pair = diagnose_decimal_point_error(left_value, right_value, expr)
            if dec_pair:
                process_score = 0
                reason = f"{dec_pair}。{correct_steps}".strip()
            elif is_decimal_mul_context(expr, student_work, reason):
                process_score = max(min(process_score, 6), 4)
                classified = classify_arithmetic_error_reason(
                    expr, student_work, student_answer, expected_answer, reason
                )
                reason = f"{classified or '【计算结果错误】末步计算有误'}。{correct_steps}".strip()
            elif is_last_step_only_arithmetic_error(expr, student_work, invalid_steps):
                process_score = max(min(process_score, 6), 4)
                reason = (
                    f"【过程正确、结果错误】末步计算有误：{left}={format_answer_display(left_value, expr)}，"
                    f"但 {right}={format_answer_display(right_value, expr)}；标准答案为 "
                    f"{format_answer_display(standard_value, expr)}。"
                ).strip()
            else:
                process_score = 0 if first_step_wrong else min(process_score, 3)
                reason = (
                    f"{reason} 说明：过程等号不成立，{left}={format_answer_display(left_value, expr)}，"
                    f"但 {right}={format_answer_display(right_value, expr)}，不能判为正确。{correct_steps}"
                ).strip()
        elif (
            final_standard is not None
            and answer_value is not None
            and not matches_any_standard(
                answer_value,
                collect_comparable_standards(resolved, expected_answer) or [final_standard],
            )
        ):
            status = "错误"
            result_score = 0
            process_score = min(process_score, 0 if first_step_wrong else 5)
            q["expected_answer"] = format_answer_display(final_standard, expr)
            dec_hint = diagnose_decimal_point_error(final_standard, answer_value, expr)
            if dec_hint:
                reason = f"{dec_hint}。{correct_steps}".strip()
            elif is_decimal_mul_context(expr, student_work, reason):
                classified = classify_arithmetic_error_reason(
                    expr, student_work, student_answer, expected_answer, reason
                )
                reason = (
                    f"{classified or '【计算结果错误】运算结果与标准答案不一致'}。{correct_steps}"
                ).strip()
            else:
                reason = (
                    f"{reason} 说明：题目数值结果应为 {q['expected_answer']}，"
                    f"与当前作答不一致。{correct_steps}"
                ).strip()
        elif standard_value is not None and student_work and not arith_chain_ok:
            status = "错误"
            result_score = 0
            process_score = min(process_score, 0 if first_step_wrong else 5)
            q["expected_answer"] = q.get("expected_answer") or format_answer_display(standard_value, expr)
            if is_decimal_mul_context(expr, student_work, reason):
                classified = classify_arithmetic_error_reason(
                    expr, student_work, student_answer, expected_answer, reason
                )
                reason = f"{classified or '【计算结果错误】运算结果与标准答案不一致'}。{correct_steps}".strip()
            else:
                reason = (
                    f"{reason} 说明：按题干重算标准结果约为 {q['expected_answer']}，与作答过程或最终结果不一致。{correct_steps}"
                ).strip()

    return status, process_score, result_score, structure_score, reason


def normalize_question(q: dict) -> dict:
    """兜底规范 AI 返回，避免缺字段或误把未作答/过程错结果对判成正确。"""
    q = q or {}
    patched = patch_contradictory_model_verdict(q)
    if patched is not None:
        return patched
    status = str(q.get("status", "错误")).strip()
    student_work_raw = strip_invalidated_work(str(q.get("student_work", q.get("work", "")) or "").strip())
    student_answer_raw = strip_invalidated_work(str(q.get("student_answer", q.get("answer", "")) or "").strip())
    expr = str(q.get("expr", "")).strip()

    if lacks_effective_student_response(expr, student_work_raw, student_answer_raw):
        q["expr"] = expr
        q["student_work"] = ""
        q["student_answer"] = ""
        q["status"] = "未作答"
        q["process_score"] = 0
        q["result_score"] = 0
        q["structure_score"] = 0
        q["total_score"] = 0
        q["reason"] = sanitize_reason_text(
            "题目有作答区域，但未识别到有效脱式、计算步骤或最终答案，按未作答处理。"
        )
        return q

    expr = maybe_fix_expr_from_student_numeric(expr, student_work_raw, student_answer_raw)
    reason = remove_invalidated_reason_fragments(str(q.get("reason", "")).strip())
    expr = maybe_fix_expr_missing_percent(expr, student_work_raw, reason)
    q["expr"] = expr
    student_work, student_answer = apply_decimal_ocr_correction(
        expr,
        student_work_raw,
        student_answer_raw,
        q.get("expected_answer", ""),
        reason,
    )
    symbolic_expr = uses_non_arithmetic_grading(expr, student_work)

    if not symbolic_expr and detect_expression_structure_violation(expr, student_work):
        correct_steps = build_correct_steps(expr, student_work, reason)
        q["status"] = "错误"
        q["process_score"] = 0
        q["result_score"] = 0
        q["structure_score"] = 0
        q["total_score"] = 0
        q["student_work"] = student_work
        q["student_answer"] = student_answer
        order_msg = (
            "【运算顺序错误】括号或乘除顺序与题目不一致。"
            if is_decimal_mul_context(expr, student_work)
            else "【解题逻辑错误】括号位置或运算顺序与题目不一致（例如把 a×(b+c) 写成 (a+b)×c），"
            "属于严重列式错误，过程分、结果分均为 0。"
        )
        q["reason"] = sanitize_reason_text(f"{order_msg} {correct_steps}")
        return q
    resolved_early = resolve_grading_standard(
        expr, student_work, student_answer, q.get("expected_answer", ""), reason
    )
    if resolved_early["final"] is not None and not uses_non_arithmetic_grading(expr, student_work):
        q["expected_answer"] = format_answer_display(resolved_early["final"], expr)
    elif is_radical_expression_context(expr, student_work):
        rad_std = eval_radical_expression(expr)
        if rad_std is not None:
            q["expected_answer"] = format_radical_answer(rad_std)
    expr_truth = resolved_early["exact"] or (safe_eval_math_expr(expr) if expr else None)
    final_truth = resolved_early["final"] or expr_truth
    answer_only_correct = bool(
        student_answer_raw
        and (
            student_answer_matches_expr(expr, student_work, student_answer_raw)
            or (
                final_truth is not None
                and extract_math_value(student_answer_raw) is not None
                and values_equal(extract_math_value(student_answer_raw), final_truth)
            )
        )
    )
    arithmetically_correct = student_solution_arithmetically_correct(
        expr, student_work, student_answer, q.get("expected_answer", ""), reason
    ) or answer_only_correct
    rounding_compliant = student_approximate_rounding_correct(
        expr, student_work, student_answer, reason
    ) or (
        reason_claims_compliant_rounding(reason)
        and (
            student_answer_matches_expr(expr, student_work, student_answer)
            or bool(numeric_outcome_matches(expr, student_work, student_answer, q.get("expected_answer", ""), reason).get("match"))
        )
    )
    radical_reason_ok = (
        is_radical_expression_context(expr, student_work)
        and reason_claims_student_fully_correct(reason)
        and (
            bool(str(student_work or "").strip() or str(student_answer or "").strip())
            or reason_has_contradictory_false_negative(reason)
        )
    )
    contradiction_correct = (
        status == "错误"
        and (
            arithmetically_correct
            or rounding_compliant
            or radical_reason_ok
            or radical_solution_correct(
                expr, student_work, student_answer, q.get("expected_answer", ""), reason
            )
        )
        and not looks_like_only_copied_question(expr, student_work)
        and not first_step_invalid(expr, student_work, student_answer, q.get("expected_answer", ""))
        and not detect_expression_structure_violation(expr, student_work)
    )
    if contradiction_correct:
        status = "正确"
        q["process_score"] = max(clamp_score(q.get("process_score"), 0, 6), 6)
        q["result_score"] = max(clamp_score(q.get("result_score"), 0, 3), 3)
        q["structure_score"] = max(clamp_score(q.get("structure_score"), 0, 1), 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    no_answer_markers = ("未作答", "空白", "没写", "无作答", "未填写", "没有作答", "答题区为空", "没有有效作答")
    blank_like_answers = ("", "无", "空", "空白", "未写", "未作答", "没写", "无作答", "没有")
    has_no_effective_answer = student_work in blank_like_answers and student_answer in blank_like_answers
    wants_unanswered = status == "未作答" or (
        has_no_effective_answer and (not reason or any(m in reason for m in no_answer_markers))
    )
    if wants_unanswered:
        if lacks_effective_student_response(expr, student_work, student_answer):
            status = "未作答"
            q["student_work"] = ""
            q["student_answer"] = ""
            q["process_score"] = 0
            q["result_score"] = 0
            q["structure_score"] = 0
            q["total_score"] = 0
            q["reason"] = sanitize_reason_text(reason or "题目有作答区域但没有看到有效书写，按未作答处理。")
            return q
        status = "错误"
        reason = (reason or "卷面旁有手写答案，已按识别到的数值批改，不再按未作答处理。").strip()

    process_score = clamp_score(q.get("process_score"), 0, 6)
    result_score = clamp_score(q.get("result_score"), 0, 3)
    structure_score = clamp_score(q.get("structure_score"), 0, 1)

    error_keywords = (
        "过程错", "过程错误", "中间错", "等号不成立", "抄错", "列式错", "符号错", "括号",
        "进位错", "退位错", "借位错", "通分错", "约分错", "漏项", "单位错", "计算错",
        "94×20", "1880", "碰巧", "明显错误", "严重错误"
    )
    if status == "过程不规范" and any(k in reason for k in error_keywords) and not arithmetically_correct and not symbolic_expr:
        status = "错误"
        process_score = min(process_score, 3)
        if not reason.endswith("按错误处理。"):
            reason = f"{reason} 过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。"

    status, process_score, result_score, structure_score, reason = local_arithmetic_guard(
        q, status, process_score, result_score, structure_score, reason
    )

    arith_after_guard = student_solution_arithmetically_correct(
        expr, student_work, student_answer, q.get("expected_answer", ""), reason
    )

    if (
        arith_after_guard
        and status == "错误"
        and (
            final_answer_acceptable(expr, student_work, student_answer, q.get("expected_answer", ""), reason)
            or not detect_invalid_equal_steps(student_work, expr, student_work, reason)
        )
        and not looks_like_only_copied_question(expr, student_work)
        and not first_step_invalid(expr, student_work, student_answer, q.get("expected_answer", ""))
        and not detect_expression_structure_violation(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    if (
        status == "错误"
        and rounding_compliant
        and not detect_expression_structure_violation(expr, student_work)
        and not looks_like_only_copied_question(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = strip_false_negative_grading_hints(reason) or (
            "先算精确乘积，再按题意四舍五入保留小数位数，作答正确。"
        )

    if (
        status == "错误"
        and is_numeric_eval_algebra_task(expr, student_work, q.get("expected_answer", ""))
        and algebra_numeric_outcome_ok(expr, student_work, student_answer, q.get("expected_answer", ""))
        and not first_step_invalid(expr, student_work)
        and not looks_like_only_copied_question(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    if (
        status != "未作答"
        and student_work
        and first_step_invalid(expr, student_work, student_answer, q.get("expected_answer", ""))
        and str(status).strip() != "正确"
        and not arith_after_guard
        and not reason_claims_student_fully_correct(reason)
    ):
        if opposite_sign_final_only(expr, student_answer, student_work):
            status = "错误"
            process_score = max(process_score, 4)
            result_score = 0
            reason = re.sub(r"\s*提示：第一步已从原式错误变形[^。]*。", "", reason).strip()
        else:
            status = "错误"
            process_score = 0
            result_score = 0
            structure_score = 0
            if "第一步" not in reason and "从原式错误变形" not in reason:
                reason = (reason + " 提示：第一步已从原式错误变形，后续步骤不再计过程分或结果分。").strip()

    if status == "错误" and result_score == 0 and process_score >= 5:
        reason_text = reason or "过程基本正确但最终答案错误。"
        if any(k in reason_text for k in ("结果", "答案", "最终", "计算细节")) and not any(k in reason_text for k in error_keywords):
            process_score = min(process_score, 6)

    status, process_score, result_score, reason = apply_process_correct_result_wrong(
        expr, student_work, student_answer, status, process_score, result_score, reason
    )

    if (
        status == "错误"
        and (
            student_approximate_rounding_correct(expr, student_work, student_answer, reason)
            or (
                reason_claims_compliant_rounding(reason)
                and numeric_outcome_matches(
                    expr, student_work, student_answer, q.get("expected_answer", ""), reason
                ).get("match")
            )
        )
        and not detect_expression_structure_violation(expr, student_work)
        and not looks_like_only_copied_question(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = strip_false_negative_grading_hints(reason) or (
            "先算精确乘积，再按题意四舍五入保留小数位数，作答正确。"
        )

    # 方程/配方类：仅书写格式、未写解集/检验、下标写法等，不单独构成「过程不规范」
    if _equation_or_algebra_hint(expr, student_work) and status == "过程不规范":
        if not any(k in reason for k in error_keywords) and result_score >= 2 and process_score >= 4:
            status = "正确"
            process_score = max(process_score, 6)
            result_score = max(result_score, 3)
            structure_score = max(structure_score, 1)

    calculated_total = process_score + result_score + structure_score
    total_score = clamp_score(calculated_total, 0, 10)
    if status == "正确" and total_score < 9:
        if is_radical_expression_context(expr, student_work) and (
            student_solution_arithmetically_correct(
                expr, student_work, student_answer, q.get("expected_answer", ""), reason
            )
            or reason_claims_student_fully_correct(reason)
        ):
            process_score = max(process_score, 6)
            result_score = max(result_score, 3)
            structure_score = max(structure_score, 1)
            total_score = 10
        else:
            status = "过程不规范" if total_score >= 7 else "错误"
    if status == "错误" and total_score > 8:
        total_score = 8
    if (
        status in ("错误", "未作答")
        and expr
        and "正确步骤" not in reason
        and not symbolic_expr
        and not is_decimal_mul_context(expr, student_work, reason)
    ):
        reason = f"{reason} {build_correct_steps(expr, student_work, reason)}".strip()

    if str(status).strip() == "正确":
        reason = strip_false_negative_grading_hints(reason)

    final_arith_ok = (
        student_solution_arithmetically_correct(
            expr, student_work, student_answer, q.get("expected_answer", ""), reason
        )
        or answer_only_correct
    ) and not detect_expression_structure_violation(expr, student_work)
    status, reason = reconcile_reason_with_verdict(reason, status, final_arith_ok, student_work, expr)
    if status == "正确" and final_arith_ok:
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        total_score = clamp_score(process_score + result_score + structure_score, 0, 10)

    if is_decimal_mul_context(expr, student_work, reason) and not symbolic_expr:
        resolved_final = resolve_grading_standard(
            expr, student_work, student_answer, q.get("expected_answer", ""), reason
        )
        if resolved_final["final"] is not None:
            q["expected_answer"] = format_answer_display(resolved_final["final"], expr)
        if str(status).strip() == "正确":
            reason = ensure_decimal_only_presentation(
                strip_false_negative_grading_hints(reason)
                or "先算精确乘积，再按题意四舍五入保留小数位数，作答正确。",
                expr,
                student_work,
            )
        else:
            reason = consolidate_decimal_correction_reason(
                expr,
                student_work,
                student_answer,
                status,
                reason,
                q.get("expected_answer", ""),
            )

    status, process_score, result_score, structure_score, reason = enforce_verdict_reason_consistency(
        expr,
        student_work,
        student_answer,
        status,
        process_score,
        result_score,
        structure_score,
        reason,
        q.get("expected_answer", ""),
    )
    total_score = clamp_score(process_score + result_score + structure_score, 0, 10)

    q["expr"] = expr
    q["student_work"] = student_work
    q["student_answer"] = student_answer
    q["status"] = status if status in ("正确", "过程不规范", "错误", "未作答") else "错误"
    q["process_score"] = process_score
    q["result_score"] = result_score
    q["structure_score"] = structure_score
    q["total_score"] = total_score
    q["reason"] = sanitize_reason_text(reason or "已按题目、过程和最终答案综合核验。")
    return q


def normalize_math_result(data: dict) -> dict:
    if not data:
        return data
    questions = [normalize_question(q) for q in data.get("questions", [])]
    for i, q in enumerate(questions):
        again = patch_contradictory_model_verdict(q)
        if again is not None:
            questions[i] = again
    data["questions"] = questions
    statuses = [q.get("status") for q in questions]
    if not questions or all(s == "未作答" for s in statuses):
        data["overall_result"] = "未作答"
    elif all(s == "正确" for s in statuses):
        data["overall_result"] = "全部正确"
    elif all(s in ("错误", "未作答") for s in statuses):
        data["overall_result"] = "全部错误"
    else:
        data["overall_result"] = "部分正确"
    return data


def _teacher_context_suffix_math(
    grade_level: str,
    teacher_note: str,
    answer_key: str = "",
    scoring_rubric: str = "",
    *,
    has_answer_key_image: bool = False,
) -> str:
    g = (grade_level or "").strip()
    n = (teacher_note or "").strip()
    ak = (answer_key or "").strip()
    rub = (scoring_rubric or "").strip()
    if not g and not n and not ak and not rub and not has_answer_key_image:
        return ""
    parts = ["\n\n【任课教师补充说明（请纳入批改语境；若与卷面或题干明显冲突以卷面与题干为准）】"]
    if has_answer_key_image:
        parts.append(
            "【参考答案】已在前一张附图中给出，请对照该图核对学生各题作答与过程，提高判分一致性与准确率；"
            "卷面看不清或学生未作答时仍以卷面可见内容为准，勿臆造学生步骤。"
        )
    elif ak:
        parts.append(
            "【教师提供的参考答案（选填，请作为判分重要依据）】\n"
            f"{ak}\n"
            "请对照参考答案核对学生各题作答与过程，提高判分一致性与准确率；"
            "若卷面与参考答案在书写上等价（如分数/小数/单位等价）应判对；"
            "卷面看不清或学生未作答时仍以卷面可见内容为准，勿臆造学生步骤。"
        )
    if rub:
        parts.append(
            f"【评分细则】\n{rub}\n"
            "请在过程分与结果分上尽量按细则把握。"
        )
    if g:
        parts.append(
            f"学生所在年级（老师填写）：{g}。请按该学段课标与常见认知水平把握判分尺度、总评语气与 reason 详略；不要超纲使用明显超出该学段要求的术语硬套学生。"
        )
    if n:
        parts.append(
            f"老师备注 / 学情说明：{n}\n请结合以上内容辅助判断掌握程度与常见问题，并在总评、薄弱点或书写习惯建议中酌情呼应（无需重复摘抄备注全文）。"
        )
    return "\n".join(parts)


def call_ai_math(
    image_path: str,
    *,
    grade_level: str = "",
    teacher_note: str = "",
    answer_key: str = "",
    answer_key_image: str = "",
    scoring_rubric: str = "",
) -> dict:
    try:
        from core.moonshot_client import moonshot_configured, moonshot_vision_chat
    except ImportError:
        moonshot_configured = lambda: False  # type: ignore
        moonshot_vision_chat = None  # type: ignore

    use_moonshot = moonshot_configured()
    if not use_moonshot and dashscope is None:
        logging.error("未配置 MOONSHOT_API_KEY，且 dashscope 未安装，无法调用视觉批改 API")
        return None
    try:
        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        image_uri = f"data:{get_mime_type(image_path)};base64,{b64}"

        suffix = _teacher_context_suffix_math(
            grade_level,
            teacher_note,
            answer_key,
            scoring_rubric,
            has_answer_key_image=bool((answer_key_image or "").strip() and os.path.isfile(answer_key_image)),
        )
        content: list[dict] = []
        ak_path = (answer_key_image or "").strip()
        if ak_path and os.path.isfile(ak_path):
            with open(ak_path, "rb") as f:
                ak_b64 = base64.b64encode(f.read()).decode()
            content.append({"image": f"data:{get_mime_type(ak_path)};base64,{ak_b64}"})
            content.append({"text": "【上图】为教师提供的参考答案，请先阅读再批改下一张学生作业。"})
        content.append({"image": image_uri})
        content.append({"text": MATH_PROMPT + suffix})
        messages = [{"role": "user", "content": content}]

        raw: str | None = None
        if use_moonshot and moonshot_vision_chat is not None:
            raw, err = moonshot_vision_chat(
                content,
                temperature=0.05,
                max_tokens=8192,
                timeout=120,
            )
            if err:
                logging.error(f"Kimi 批改失败：{err}")
                return None
        else:
            # 兼容：未配置 Moonshot 时回退 DashScope
            resp = dashscope.MultiModalConversation.call(
                model=os.getenv("DASHSCOPE_VL_MODEL", "qwen-vl-max"),
                messages=messages,
                timeout=120,
                max_length=8192,
                temperature=0.05,
                top_p=0.3,
            )
            if resp.status_code == HTTPStatus.OK:
                raw = resp.output.choices[0].message.content[0]["text"]
            else:
                logging.error(f"DashScope API 失败：{resp.status_code} | {resp.message}")
                return None

        if raw:
            try:
                return normalize_math_result(json.loads(extract_json_safe(raw)))
            except Exception as norm_err:
                logging.error(f"结果规范化失败：{norm_err}", exc_info=True)
                return None
        return None
    except Exception as e:
        logging.error(f"处理异常：{e}", exc_info=True)
        return None


# ✅ 修复点 1：参数名改为 data，if 补全 data
def format_console_output(data: dict, filename: str):
    if not data:
        print(f"{Fore.RED}❌ 批改失败\n")
        return

    status = data.get("overall_result", "未知")
    status_map = {
        "全部正确": (Fore.GREEN, "✅ 全部正确"),
        "部分正确": (Fore.YELLOW, "⚠️ 部分正确"),
        "全部错误": (Fore.RED, "❌ 全部错误"),
        "未作答": (Fore.BLUE, "📝 未作答")
    }
    color, status_text = status_map.get(status, (Fore.WHITE, status))

    print(f"\n{Fore.CYAN}{'=' * 70}")
    print(f"📄 文件：{filename}")
    print(f"🏷️  状态：{color}{status_text}{Style.RESET_ALL}")

    print(f"\n📝 逐题详情:")
    for q in data.get("questions", []):
        q_status = q.get("status", "")
        q_color = Fore.GREEN if q_status == "正确" else Fore.YELLOW if q_status == "过程不规范" else Fore.RED if q_status == "错误" else Fore.BLUE
        score_text = f"{q.get('total_score', 0):g}/10"
        print(
            f"  {q['id']}. {q_color}{prettify_math(q.get('expr', ''))}{Style.RESET_ALL} → {q_color}{q_status}{Style.RESET_ALL} ({score_text})")
        if q.get("student_work"):
            print(f"     {Style.DIM}学生过程：{prettify_math(q['student_work'])}{Style.RESET_ALL}")
        if q.get("expected_answer"):
            print(f"     {Style.DIM}标准结果：{prettify_math(q['expected_answer'])}{Style.RESET_ALL}")
        print(
            f"     {Style.DIM}得分：过程 {q.get('process_score', 0):g}/6，结果 {q.get('result_score', 0):g}/3，规范 {q.get('structure_score', 0):g}/1{Style.RESET_ALL}")
        print(f"     {Style.DIM}└─ {q.get('reason', '')}{Style.RESET_ALL}")

    print(f"\n💬 老师评语:")
    for line in textwrap.wrap(data.get("personal_comment", "暂无"), width=66, initial_indent="   ",
                              subsequent_indent="   "):
        print(line)

    print(f"\n⚠️  薄弱知识点:")
    weaks = data.get("weak_points", [])
    for w in (weaks if weaks else ["暂无显著薄弱点，基础扎实，继续保持！"]):
        print(f"   • {w}")
    print(f"{Fore.CYAN}{'=' * 70}{Style.RESET_ALL}\n")


def generate_class_analysis(results: list) -> str:
    if not results: return "无有效数据"
    weak_counter, status_counter = {}, {"全部正确": 0, "部分正确": 0, "全部错误": 0, "未作答": 0}
    for r in results:
        status_counter[r.get("overall_result", "全部错误")] += 1
        for pt in r.get("weak_points", []): weak_counter[pt] = weak_counter.get(pt, 0) + 1

    top_weak = sorted(weak_counter.items(), key=lambda x: x[1], reverse=True)[:5]
    lines = ["========================================", "📊 数学作业学情分析报告",
             "========================================",
             f"📝 总份数：{len(results)}", f"✅ 全部正确：{status_counter['全部正确']}",
             f"⚠️ 部分正确：{status_counter['部分正确']}",
             f"❌ 全部错误：{status_counter['全部错误']}", f"📝 未作答：{status_counter['未作答']}", "",
             "🔥 高频薄弱知识点 (Top 5):"]
    for k, v in top_weak: lines.append(f"   - {k}: {v} 人次")
    lines.append("\n💡 教学建议:")
    lines.append(f"   建议针对【{top_weak[0][0]}】开展专项训练。" if top_weak else "   整体掌握良好，继续保持。")
    return "\n".join(lines)


# ================= 5. 主流程 =================
def batch_correct_math():
    folder = Path(MATH_FOLDER)
    if not folder.exists():
        print(f"{Fore.RED}❌ 请创建 {MATH_FOLDER} 文件夹");
        return

    imgs = sorted(list(folder.glob("*.jpg")) + list(folder.glob("*.png")) + list(folder.glob("*.jpeg")))
    if not imgs:
        print(f"{Fore.YELLOW}⚠️ 文件夹内无图片");
        return

    print(f"{Fore.CYAN}📚 启动 AI 数学作业批改系统...\n")
    results = []

    for img in imgs:
        print(f"{Fore.YELLOW}⏳ 批改：{img.name}")
        data = call_ai_math(img)
        # ✅ 修复点 2：if 后面补全 data
        if data:
            data["filename"] = img.name
            results.append(data)
            format_console_output(data, img.name)
        else:
            print(f"{Fore.RED}   ❌ 失败\n")

    if not results:
        print(f"{Fore.RED}\n⚠️ 未成功批改任何作业");
        return

    print("\n" + generate_class_analysis(results))

    wb = Workbook()
    ws = wb.active;
    ws.title = "数学批改"
    headers = ["文件名", "总评", "平均分", "逐题详情 (JSON)", "评语", "薄弱点"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True);
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        questions = r.get("questions", [])
        avg_score = round(sum(q.get("total_score", 0) for q in questions) / len(questions), 1) if questions else 0
        ws.append([r["filename"], r["overall_result"], avg_score,
                   json.dumps(questions, ensure_ascii=False),
                   r.get("personal_comment", ""), ", ".join(r.get("weak_points", []))])
    ws.column_dimensions['D'].width = 80
    wb.save(OUTPUT_EXCEL)

    with open(OUTPUT_TXT, "w", encoding="utf-8") as f:
        f.write(generate_class_analysis(results) + "\n\n")
        for r in results:
            f.write(f"--- {r['filename']} | {r['overall_result']} ---\n")
            for q in r.get("questions", []):
                f.write(f"  {q['id']}. {prettify_math(q['expr'])} → {q['status']} ({q['reason']})\n")
            f.write(f"  评语：{r.get('personal_comment', '')}\n\n")

    print(f"\n{Fore.GREEN}✅ 完成！报告已保存至 {OUTPUT_EXCEL} 和 {OUTPUT_TXT}")


if __name__ == "__main__":
    batch_correct_math()