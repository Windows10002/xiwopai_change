import os
import re
import json
import logging
import textwrap
import base64
import sys
import subprocess
import ast
import operator
from fractions import Fraction
from decimal import Decimal, InvalidOperation
from pathlib import Path
from http import HTTPStatus
import dashscope
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ================= 1. 环境初始化 =================
if sys.platform == 'win32':
    try:
        subprocess.run(['chcp', '65001'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

init(autoreset=True)
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

# ================= 2. 配置区 =================
dashscope.api_key = os.getenv("DASHSCOPE_API_KEY", "sk-5d1b92362694429f81c2eb03d6988e3e")
MATH_FOLDER = "img_math"
OUTPUT_EXCEL = "数学作业批改报告.xlsx"
OUTPUT_TXT = "数学作业批改报告.txt"

# ================= 3. 核心 Prompt =================
MATH_PROMPT = """
你是资深小学/初中数学阅卷教师。请公平、审慎批改图片中的手写作答：既要防止明显错误被放过，更要防止「推理与答案实质正确却被判错」尤其是 OCR 把下标、乘号、涂改笔画误读时。

【识别要求】
1. 必须先完整找出图片中的所有题号/横式/竖式/应用题，逐题输出，禁止遗漏；不要只输出写了答案的题。
2. 对每个题号先判断“是否有有效作答”。如果题干存在但答题区域为空、只有题号、只有抄题没有计算/答案、只有草稿痕迹、只有无关涂画、答案被划掉且没有替代答案，必须判为【未作答】，student_work 和 student_answer 均填空字符串。
2b. **questions 数组必须覆盖卷面上每一道可辨认题目（含横式/竖式/脱式/应用题），题量须与卷面一致，禁止只输出前几题或省略中间题号；若一页超过 20 题仍须全部输出，不得因长度自行截断。**
2c. **仅把原题横式/算式照抄一遍、没有「=」后的有效变形或计算、或「=」两侧只是重复题干而无新数值，一律判【错误】或【未作答】，process_score、result_score、structure_score 全部为 0。**
2d. **划改识别（极其重要）**：凡被手写横线、叉号、涂改液、橡皮擦淡影明显划掉的一行或一整段演算，一律视为作废，不得写入 student_work，不得在 reason 中引用、摘抄或点评；只保留划改之后最终仍清晰可见的有效书写。若划改后另起一行重写，以重写行为准。
2e. **首行「=」右侧优先（极其重要）**：若紧接「=」的右侧已出现清晰、连贯的有效脱式（例如「30×42」或「30 × 42」），而「=」左侧仅有涂抹、杂乱笔画、被横线划掉的区域或无法辨认的乱迹，**禁止**把左侧乱迹误读成算式片段再与右侧拼接（典型误读：把乱迹+「30×42」读成「46-30×42」）。此时应认定学生已先完成括号内运算得到 30，再与 42 相乘；student_work 只收录「=」右侧起的有效链。
3. 严禁把下一题/上一题的作答挪到本道题。每题只能读取该题号右侧、下方且空间归属明确的书写；若本道题答题区为空，哪怕下一行/下一题有步骤，也要判本道题【未作答】。
4. 学生明显涂改、划掉、叉掉、擦除残影、重写前的旧过程，一律视为无效痕迹，不能写入 student_work，不能参与过程分和结果分；只批改最后未被划掉的有效作答。reason 中也不要把被划掉的错误数字/错误步骤当成错误原因。
5. 遇到一页中有多题时，必须按从上到下、从左到右顺序逐题编号；看不清但能确认题目存在时也要输出，并在 reason 中说明“题目/作答疑似模糊”。
6. OCR 不确定时，在 reason 中说明「疑似识别为…」，但仍要按可见内容批改，不能因为不确定就默认正确。
6b. **印刷体与手写易混（极其重要）**：`+` 与 `÷`、`×` 与 `x`、小数点与乘号、带分数「整数 空格 分数」与「整数乘分数」要对照印刷题干整体判断；**若印刷体明确为 (a-b)×c、(a+b)÷d 等括号结构，学生先算括号再乘除且数值链成立，不得因 expr 字段抄错运算符而改判为错。**
6c. **expr 字段**必须与卷面印刷题干一致；若题干含括号而你写的 expr 丢了括号或把 × 写成 +，应先按图片修正 expr 再批改，避免后续「标准结果」与卷面不符。
6d. **除号与分数（极其重要）**：题干或脱式中出现「÷ a/b」（a、b 为整数）时，必须理解为「÷ (a/b)」，即除以整个分数 **a/b**，**禁止**理解为「(÷ a) ÷ b」或从左到右拆成 `/a/b` 导致数值错误。同级 **× 与 ÷** 仍按课标从左到右依次计算；**禁止**先做后面的乘法再做前面的除法。
6d2. **带分数与「+1/3」易混（极其重要）**：印刷体在「-1/6 - 1/4 + ___」这类括号末项若是 **带分数 1 1/3**，整数与分数之间常有空隙，切勿误读为单独的 **+1/3**；expr 必须按卷面带分数抄写（写作 `1 1/3`），否则整题标准值会错。
6d3. **带分数与假分数书写等价（极其重要）**：学生在括号内写 **1 1/3** 或与它数值等价的 **4/3、16/12** 等通分形式，只要等号链每步成立、结果与按题干重算一致，**不得**声称「把 1 1/3 写成假分数所以通分错」；**禁止**把真分数 **11/12** 误写成带分数「4又 1/12」等离奇形式来否定学生；更不得把学生已算对的 **144/11** 说成与标准答案不一致。
6e. **expected_answer** 必须与按 expr 在实数范围内重算的结果一致（允许带分数、小数、最简分数等书写等价）；若你写的 expected_answer 与重算不符，以重算为准，不得据此把学生判错。

【计算核验规则】
1. 对每道计算题必须独立重算标准结果，再核对学生过程和最终答案；先验算题干原式，再验算每一个等号右侧是否由左侧正确推出。
2. “按步骤给分”只奖励错误发生前的正确步骤：如果第一步就抄题、变形、运算顺序、括号处理、符号或等号关系错误，process_score 必须为 0；不能因为后面继续算对了错误式子而给过程分。
2b. **脱式首行省略（极其重要）**：若学生从中间某步写起（如仅「=976÷8=…」而前面「(608+368)÷8」在卷面上行已写清），只要该步数值与题干在该阶段应得值一致，**不得**把「第一个等号」误判为从原式第一步就错；仅最后一步算术抄错时，须保留过程分（通常 4～6），result_score=0。
3. 对「纯数字四则/脱式」：中间过程出现等号不成立、运算对象抄错、括号/符号/顺序错等，即使最终答案碰巧与正确答案相同，也判【错误】。对「方程、配方、因式分解、开平方得±、多根」等含未知数的题：不得以「相邻两步之间没有用纯数字验算」为由判整条错误；只要配方/开方/求根思路正确且求得的根代入能满足方程，应判【正确】或【过程不规范】，不得判【错误】。
3b. **末步算术错误（极其重要）**：若前几步括号与变形均正确，仅最后一步加减乘除算错，使 student_answer 与 expected_answer 数值不一致，**禁止**判【正确】；必须判【错误】，result_score=0，process_score 通常给 4～6 表示「前半过程正确」。
3c. **仅符号相反（极其重要）**：括号内通分、除法变乘法等前几步均正确，仅最后一步把「÷(-1/12)」写成「×12」等导致结果为正、标准值为负（或反之），**必须**保留过程分（通常 4～6），result_score=0，**禁止**套用「第一步已从原式错误变形」清零；reason 应点明「符号错误」。
3d. **多步小数与带分数混合**：若学生将可精确抵消的项（如 -5.25 与 +5¼）先行抵消，再把余下的带分数与整数近似合并书写（如写作「+6」），只要脱式等号链在**其书写下自洽**且最终分式与按题干严格通分结果相差不超过 **1/4**（初中练习常见取整/近似容差），判【正确】或【过程不规范】，不得仅因与冗长通分式数值相差极小而判【错误】。
3d2. **跳步简便运算（极其重要）**：学生首行未逐字抄全原式，但已将若干项**合理合并**（如抵消小数与带分数、先合并同分母分数），只要**每一个等号两侧数值成立**且**最终结果与按题干重算一致**，必须判【正确】或【过程不规范】，**禁止**套用「第一步已从原式错误变形」清零，也**禁止**在订正提示中写「末步与标准答案不一致」类自相矛盾表述。
3e. **先化简再求值**：若化简式与代入后的数值答案与 expected_answer 一致，即使 reason 中挑剔中间某行笔误，只要 student_answer 正确，必须判【正确】或【过程不规范】，不得判【错误】。
4. 如果过程描述为“过程正确、每步成立、结果正确”，status 必须是【正确】或【过程不规范】，不得同时判为【错误】；除非明确指出哪一步具体错误。
5. 示例：(52-7)×6 写成 =52-(7×6) 是第一步错误，process_score=0，标准步骤为 (52-7)×6=45×6=270。
6. 示例：979÷(144-133) 如果本道题无有效书写，不得把下一题的 (16+34)×30 识别成本题作答；应判【未作答】。
7. 示例：21×(60-55)=21×5=105 每一步都成立，必须判【正确】，不能因模板化语句写成错误。
8. 示例：22×(16+21) 中被涂改/划掉的错误过程要忽略，只看最后保留的有效过程；如果最后有效过程正确，应判正确。再例如学生先写了错误数字又划掉，后面改成正确答案，划掉内容不得出现在 student_work 和 reason 中。
8b. 示例：(46-16)×42 卷面若在第一个「=」左侧有重涂/涂改，只要「=」后有效书写为「30×42 = 1260」且每步数值成立，必须按「先算括号得 30 再乘 42」理解并优先判【正确】；**除非**在未被划掉的区域真的写出了「46-30×42」这类整段式子，否则不得在 reason 中声称学生写成了「46-30×42」或「违反运算顺序」。
9. 常见题型都要严查：整数/小数/分数四则混合运算、脱式计算、简便计算、竖式计算、解方程、比例、百分数、单位换算、应用题列式、几何周长面积体积、带余除法。
10. 允许合理等价形式：分数等价、约分后相等、乘除互化、交换律/结合律/分配律正确使用；但“等价”必须每步成立。
11. 竖式题要核对每一位、进位/退位/小数点位置；应用题要核对数量关系、列式、单位和答句；方程题要核对移项/合并/两边同乘除/检验。

【评分权重，满分 10 分】
- process_score 过程分 6 分：列式/方法 2 分，关键步骤与运算顺序 2 分，每步计算细节 2 分。
- result_score 结果分 3 分：数值正确 2 分，最终形式/化简/单位正确 1 分。
- structure_score 规范分 1 分：书写格式、答句、单位、约分、竖式对齐等规范。解方程时 x1/x2 与 x₁/x₂、未单独写「解集」或检验步骤，**不**视为格式错误，不得因此把【正确】降为【过程不规范】或扣光 structure_score。
- 过程正确但结果错：process_score 通常给 4-6，result_score 必须为 0，structure_score 按书写给 0-1，status 为【错误】。
- 过程基本对但有轻微算术失误：process_score 通常给 3-5，result_score 为 0，status 为【错误】。
- 「纯算术」结果正确但脱式中等号链不成立/抄题错：status 必须为【错误】；如果第一步就错，process_score=0；如果中途才错，process_score 只给错误前已正确步骤，通常不超过 3；result_score 最多给 1-2，不能因答案碰巧相同给满分。含未知数的方程/配方/开方题不适用「逐段纯数字验等号」的严苛标准，见上【计算核验规则】第 3 条。
- 只有答案无过程：若题目要求过程/脱式/竖式/应用题，process_score 不超过 2；若答案正确可给 result_score，status 通常为【过程不规范】。
- 未作答：三项均为 0，status 为【未作答】。
- 被涂改/划掉的旧过程：不计入 student_work，不给分也不扣分，只按最终保留的有效作答评分。
- 错误或未作答题目必须在 reason 中给出注解，并写出完整正确步骤及答案。

【订正说明 reason 的书写规范（面向学生展示）】
1. 只用「老师批改学生作业」的口吻，短句、条理清楚；禁止出现「模型/AI/大模型/OCR/识别器/置信度/猜测/分析过程」等字样，禁止自我辩解或复述你的内心推理。
2. **正确**或**过程不规范**题：用一两句肯定写对之处即可，不必写长篇「解析」。
3. **错误**题：reason 必须用四段（每段一行起首，便于展示），格式严格如下（无内容可写「无」）：
   【错误原因】……
   【正确步骤】……
   【正确答案】……
   【薄弱点】……
4. **正确**或**过程不规范**题：一两句肯定即可；不要写四段模板。
5. 化简/整式题：按同类项分组说明（如含 ab、a³b²、常数项分别合并），不要漏项；不要凭直觉给答案。
6. 纯算术题若学生最终结果与标准答案不同，必须判【错误】；若过程不成立即使结果碰巧相同也必须判【错误】。

【状态标准】
- 【正确】：过程、结果、规范都正确，总分通常 9-10。
- 【过程不规范】：核心方法、变形与最终数值/根正确，仅有轻微跳步、未写检验/解集集合形式、下标写成 x1/x2 与 x₁/x₂ 混用、书写略乱等，总分通常 7-9；不得仅因此类格式问题判【错误】。
- 【错误】：答案错、关键变形错、抄题错、符号错、计算错、逻辑错，或「纯算术题」中结果碰巧对但过程不成立。
- 【未作答】：题目存在但没有有效作答。

【输出要求】
1. 输出必须为标准 JSON，禁止 Markdown 代码块或额外文字。
2. 算式统一用纯文本：分数写为 1/3，带分数写为 1 1/3，指数写为 x^2。

【JSON 格式】
{
  "overall_result": "全部正确/部分正确/全部错误/未作答",
  "questions": [
    {
      "id": 1,
      "expr": "题目原式",
      "student_work": "学生可见作答/过程；未作答则为空字符串",
      "expected_answer": "标准结果",
      "student_answer": "学生最终答案；没有则为空字符串",
      "status": "正确/过程不规范/错误/未作答",
      "process_score": 0,
      "result_score": 0,
      "structure_score": 0,
      "total_score": 0,
      "reason": "大白话说明：哪里对、哪里错、标准算法是什么"
    }
  ],
  "personal_comment": "约 80～120 字：先总评对错分布与运算能力；再各用一两句点评书写整洁度、审题细心程度、验算习惯；最后点出最值得巩固的一条建议。语气亲切、具体、可执行。",
  "weak_points": ["知识点 1", "知识点 2"]
}
"""


# ================= 4. 核心工具函数 =================

def extract_json_safe(text: str) -> str:
    block = re.search(r'```json\s*(.*?)\s*```', text, re.DOTALL)
    if block: return block.group(1)
    start, end = text.find('{'), text.rfind('}')
    if start != -1 and end != -1 and end > start:
        return text[start: end + 1]
    raise ValueError("无法定位 JSON")


def prettify_math(text: str) -> str:
    if not text: return ""
    sup_map = str.maketrans("0123456789-", "⁰¹²³⁴⁵⁶⁷⁸⁹⁻")
    text = re.sub(r'\^([\d-]+)', lambda m: m.group(1).translate(sup_map), text)
    text = text.replace('/', '⁄')
    return text


def get_mime_type(path: str) -> str:
    ext = Path(path).suffix.lower()
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "bmp": "image/bmp",
            "webp": "image/webp"}.get(ext, "image/jpeg")


def clamp_score(value, min_value=0, max_value=10):
    try:
        value = float(value)
    except (TypeError, ValueError):
        value = 0
    return max(min_value, min(max_value, value))


_ALLOWED_AST_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def _safe_eval_ast(node):
    if isinstance(node, ast.Expression):
        return _safe_eval_ast(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return Fraction(str(node.value))
    if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_AST_OPS:
        left = _safe_eval_ast(node.left)
        right = _safe_eval_ast(node.right)
        if right == 0 and isinstance(node.op, (ast.Div, ast.FloorDiv, ast.Mod)):
            raise ZeroDivisionError
        return _ALLOWED_AST_OPS[type(node.op)](left, right)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_AST_OPS:
        return _ALLOWED_AST_OPS[type(node.op)](_safe_eval_ast(node.operand))
    raise ValueError("unsupported expression")


def normalize_math_expr(expr: str) -> str:
    """将常见手写/中文数学符号转成可安全计算的 Python 表达式。"""
    if not expr:
        return ""
    text = str(expr)
    text = text.replace("（", "(").replace("）", ")").replace("[", "(").replace("]", ")")
    # 仅把显式乘号换成 *，不要把字母 x 换成 *，否则会破坏方程、配方、因式分解等含 x 的书写。
    text = text.replace("×", "*").replace("·", "*").replace("÷", "/").replace("：", "/")
    text = text.replace("－", "-").replace("—", "-").replace("–", "-").replace("＋", "+")
    text = text.replace("＝", "=").replace("≈", "=").replace("≒", "=")
    text = re.sub(r"(?<=\d),(?=\d{3}(\D|$))", "", text)
    text = re.sub(r"[^0-9+\-*/().=\s]", "", text)
    return text.strip()


def rewrite_div_fraction_pairs(text: str) -> str:
    """将 ÷ 后紧跟的「整数/整数」视为除以该分数整体，避免求值成 (((a)/b)/c) 的课标外误读。"""
    if not text or "÷" not in text:
        return str(text)
    s = str(text)
    pattern = re.compile(r"÷\s*(-?\d+)\s*/\s*(-?\d+)")
    for _ in range(64):
        m = pattern.search(s)
        if not m:
            break
        repl = f"/(({m.group(1)})/({m.group(2)}))"
        s = s[: m.start()] + repl + s[m.end() :]
    return s


def maybe_fix_expr_from_student_numeric(expr: str, student_work: str, student_answer: str) -> str:
    """当模型把常见「1 1/3」误识成「1/3」等导致 expr 与卷面不符时，用学生最终数值反推可自洽的题干变体（仅少数安全模式）。"""
    raw = str(expr or "").strip()
    if not raw:
        return raw
    raw = raw.replace("⁄", "/")
    sv = extract_math_value(student_answer) or extract_math_value(student_work)
    if sv is None and student_work and "=" in str(student_work):
        steps_tail = split_equation_steps(str(student_work))
        if steps_tail:
            tail = steps_tail[-1]
            sv = safe_eval_math_expr(tail) or extract_math_value(tail)
    if sv is None:
        return raw
    base = safe_eval_math_expr(raw)
    if base is not None and values_equal(base, sv):
        return raw

    candidates = []
    # 括号内「+ 1/3」误少整数 1：常见于 12÷(-1/6-1/4+1 1/3) 被读成 +1/3
    for m in re.finditer(r"\+\s*1\s*/\s*3(?![\d/.])", raw):
        s = raw[: m.start()] + "+1 1/3" + raw[m.end() :]
        if s != raw:
            candidates.append(s)
    # 去重保持顺序
    seen = set()
    uniq = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    for cand in uniq:
        v = safe_eval_math_expr(cand)
        if v is not None and values_equal(v, sv):
            return cand
    return raw


def convert_mixed_fractions(text: str) -> str:
    """把手写常见的「带分数」如 1 1/3、-32 1/3 转为 (4)/3、(-97)/3，便于后续安全求值。"""
    if not text:
        return ""
    s = str(text).replace("（", "(").replace("）", ")")
    pattern = re.compile(r"(?<![\d/.])(-?\d+)\s+(\d+)\s*/\s*(\d+)(?![\d/.])")

    def repl(m: re.Match) -> str:
        whole, n, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if d == 0:
            return m.group(0)
        sign = -1 if whole < 0 else 1
        wa = abs(whole)
        num = sign * (wa * d + n)
        return f"({num})/{d}"

    prev = None
    while prev != s:
        prev = s
        s = pattern.sub(repl, s)
    return s


def is_symbolic_expression(expr: str) -> bool:
    """整式化简、解方程等含字母书写：不能做纯数字等号链本地验算，也不能套用横式数值「标准答案」覆盖。"""
    if not expr or not str(expr).strip():
        return False
    raw = str(expr)
    if not re.search(r"[A-Za-z]", raw):
        return False
    compact = re.sub(r"\s+", "", raw.lower())
    if re.fullmatch(r"-?[\d.]+e[+\-]?\d+", compact):
        return False
    return True


def sanitize_reason_text(reason: str) -> str:
    """去掉面向学生的订正中不应出现的「本地复核」等元信息，压缩空白。"""
    if not reason:
        return ""
    t = str(reason)
    t = re.sub(r"本地复核[^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"复核[:：][^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"说明[:：][^。！？\n；;]*[。！？\n；;]?", "", t)
    t = re.sub(r"作为[^。！？；;]{0,40}(?:模型|AI)[^。！？；;]*[。！？；;]", "", t)
    t = re.sub(r"(?:OCR|识别器|置信度)[^。！？；;]*[。！？；;]", "", t)
    if "【错误原因】" in t or "【正确步骤】" in t:
        t = re.sub(r"[ \t\f\v]{2,}", " ", t)
        t = re.sub(r"\n{3,}", "\n\n", t).strip()
    else:
        t = re.sub(r"\s+", " ", t).strip()
        t = re.sub(r"[。！？]{2,}", "。", t)
        chunks = [c for c in t.split("正确步骤：") if c.strip()]
        if len(chunks) > 2:
            head = chunks[0].strip()
            tail = chunks[-1].strip()
            t = f"{head} 正确步骤：{tail}".strip()
    return t.strip(" ，。；")


def _equation_or_algebra_hint(expr: str, student_work: str) -> bool:
    """识别解方程、配方、多根等书写：避免被「纯算术等号链」本地校验误伤。"""
    raw = f"{expr or ''}\n{student_work or ''}"
    if not raw.strip():
        return False
    if re.search(r"(?i)(配方|因式分解|开方|方程|未知数|一元|二次|求根|两根|移项|代入|判别式)", raw):
        return True
    if "±" in raw or "∓" in raw:
        return True
    if re.search(r"(?i)\([xy]\s*[\+\-]", raw):
        return True
    if re.search(r"(?i)\bx[_\s]*[₁₂₃₄₅₆₇₈₉₀1234567890]\s*[=＝]", raw):
        return True
    if re.search(r"(?i)\bx\d\s*[=＝]", raw):
        return True
    # 多步书写且出现未知数（排除仅有 × 被误读为 x 的极少数情况：要求有 = 且像方程）
    if raw.count("=") + raw.count("＝") >= 2:
        low = raw.lower()
        if "x" in low or "y" in low:
            return True
    return False


def is_algebraic_work_context(expr: str, student_work: str) -> bool:
    """含未知数、指数、开方±、多根标注等的作答，不能用纯算术「等号链逐段求值」复核，否则会误判。"""
    raw = f"{expr or ''}\n{student_work or ''}"
    if not raw.strip():
        return False
    if _equation_or_algebra_hint(expr, student_work):
        return True
    if re.search(r"[\^⁰¹²³⁴⁵⁶⁷⁸⁹₂₃₄₅₆₇₈₉₀]", raw):
        return True
    if "±" in raw or "∓" in raw:
        return True
    if re.search(r"(?:方程|配方|因式分解|开方|未知数|一元)", raw):
        return True
    # x₁ x_1 x2（下标根）等
    if re.search(r"(?i)\bx[_\s]*[₁₂₃₀1234]|\bx\s*[=＝]", raw):
        return True
    if re.search(r"(?i)(?:\(|\s)[xyXY]\s*[+\-^=)]|[=＝]\s*[xyXY]\s*[^0-9]", raw):
        return True
    return False


def safe_eval_math_expr(expr: str):
    text = str(expr or "")
    text = text.replace("（", "(").replace("）", ")")
    text = convert_mixed_fractions(text)
    text = rewrite_div_fraction_pairs(text)
    text = normalize_math_expr(text)
    if not text or re.search(r"[=]", text):
        return None
    if not re.fullmatch(r"[0-9+\-*/().\s]+", text):
        return None
    try:
        tree = ast.parse(text, mode="eval")
        return _safe_eval_ast(tree)
    except Exception:
        return None


def values_equal(a, b, tolerance=Fraction(1, 1000000)) -> bool:
    if a is None or b is None:
        return False
    try:
        return abs(a - b) <= tolerance
    except Exception:
        return False


# 多步分数/小数混合题：合并常数时略有近似，最终与严格通分差在初中阅卷可接受范围内
_MULTISTEP_ADDSUB_SLACK = Fraction(1, 4)


def _expr_allows_numeric_slack(expr: str) -> bool:
    return len(re.findall(r"[+\-]", str(expr or ""))) >= 3


def opposite_sign_final_only(expr: str, student_answer: str, student_work: str) -> bool:
    """最终结果仅为标准值的相反数（典型：除法变乘时丢负号），不按「第一步从原式整体错」清零。"""
    if is_algebraic_work_context(expr, student_work):
        return False
    sv = safe_eval_math_expr(expr)
    av = extract_math_value(student_answer) or extract_math_value(student_work)
    if sv is None or av is None:
        return False
    return av == -sv


def algebra_numeric_outcome_ok(expr: str, student_work: str, student_answer: str, expected_answer: str) -> bool:
    """化简求值类：若最终数值与 expected 一致，则视为算术上成立（不因中间笔误描述判死）。"""
    if not is_algebraic_work_context(expr, student_work):
        return False
    if looks_like_copied_other_question(expr, student_work) or looks_like_only_copied_question(expr, student_work):
        return False
    raw = f"{expr or ''}\n{student_work or ''}"
    if not re.search(r"[\^|²³]", expr or "") and "求值" not in raw and "化简" not in (student_work or ""):
        return False
    sa = extract_math_value(student_answer) or extract_math_value(student_work)
    ea = extract_math_value(str(expected_answer or "").strip())
    if sa is None or ea is None:
        return False
    return values_equal(sa, ea)


def extract_math_value(text: str):
    """从答案文本中提取一个可比较的数值，支持整数、小数和普通分数。"""
    if not text:
        return None
    raw = convert_mixed_fractions(str(text).replace("（", "(").replace("）", ")"))
    normalized = normalize_math_expr(raw)
    if "=" not in normalized:
        ev = safe_eval_math_expr(normalized)
        if ev is not None:
            return ev
    candidates = re.findall(r"-?\d+(?:\.\d+)?\s*/\s*-?\d+(?:\.\d+)?|-?\d+(?:\.\d+)?", normalized)
    if not candidates:
        return safe_eval_math_expr(normalized)
    token = candidates[-1].replace(" ", "")
    try:
        if "/" in token:
            numerator, denominator = token.split("/", 1)
            return Fraction(Decimal(numerator)) / Fraction(Decimal(denominator))
        return Fraction(Decimal(token))
    except (InvalidOperation, ZeroDivisionError, ValueError):
        return None


def split_equation_steps(text: str):
    normalized = normalize_math_expr(text)
    if "=" not in normalized:
        return []
    return [p.strip() for p in normalized.split("=") if p.strip()]


def detect_invalid_equal_steps(text: str, expr: str = "", student_work_full: str = ""):
    """检查脱式/横式中相邻等号两侧是否数值相等（仅适用于纯算术链；代数方程勿用）。"""
    work = student_work_full or text
    if is_algebraic_work_context(expr, work):
        return []
    steps = split_equation_steps(text)
    invalid = []
    for left, right in zip(steps, steps[1:]):
        left_value = safe_eval_math_expr(left)
        right_value = safe_eval_math_expr(right)
        if left_value is not None and right_value is not None and not values_equal(left_value, right_value):
            invalid.append((left, right, left_value, right_value))
    return invalid


def strip_invalidated_work(text: str) -> str:
    """去掉模型可能识别出的涂改、划掉、作废过程，避免参与评分。"""
    if not text:
        return ""
    valid_lines = []
    invalid_markers = (
        "划掉", "划去", "涂掉", "涂改", "擦掉", "作废", "不要", "错写", "废弃",
        "×掉", "叉掉", "删除", "删去", "抹掉", "改为", "原写", "先写", "旧过程",
        "被划", "划线", "划除", "划了", "擦除", "重写前", "修改前", "无效痕迹"
    )
    keep_markers = ("最终", "最后", "改正", "改成", "保留", "未划掉", "有效")
    for raw_line in str(text).splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "改为" in line or "改成" in line:
            parts = re.split(r"改为|改成", line, maxsplit=1)
            line = parts[-1].strip(" ：:，,。") or line
        if any(marker in line for marker in invalid_markers) and not any(marker in line for marker in keep_markers):
            continue
        line = re.sub(r"（?[^，。；;]*?(?:划掉|划去|涂掉|涂改|擦掉|作废|错写|废弃|被划|划除|修改前)[^，。；;]*?）?", "", line)
        line = re.sub(r"\[[^\]]*?(?:划掉|划去|涂改|作废|无效)[^\]]*?\]", "", line).strip()
        if not line:
            continue
        if re.fullmatch(r"[xX×/\\\-—–_~≈\s]+", line):
            continue
        valid_lines.append(line)
    return "\n".join(valid_lines).strip()


def fraction_to_text(value) -> str:
    if value is None:
        return ""
    try:
        value = Fraction(value)
        return str(value.numerator) if value.denominator == 1 else f"{value.numerator}/{value.denominator}"
    except Exception:
        return str(value)


def build_correct_steps(expr: str) -> str:
    """为可计算横式生成简洁的正确步骤和答案。"""
    raw = str(expr or "")
    prepared = convert_mixed_fractions(raw.replace("（", "(").replace("）", ")"))
    standard_value = safe_eval_math_expr(prepared)
    if standard_value is None:
        return "请按题意重新列式，逐步计算并写出最终答案。"
    normalized = normalize_math_expr(prepared)
    step_lines = []
    paren_expr = normalized
    for inner in re.findall(r"\(([^()]+)\)", normalized):
        inner_value = safe_eval_math_expr(inner)
        if inner_value is not None:
            paren_expr = paren_expr.replace(f"({inner})", fraction_to_text(inner_value), 1)
    if paren_expr != normalized:
        step_lines.append(paren_expr.replace("*", "×").replace("/", "÷"))
    step_lines.append(fraction_to_text(standard_value))
    return f"正确步骤：{expr}" + "=" + "=".join(step_lines) + f"，答案：{fraction_to_text(standard_value)}。"


def first_step_invalid(expr: str, student_work: str) -> bool:
    """「第一步」错误：第一个等号两侧数值不成立，或第一步左侧整体值已与题干原式不一致。

    不要求第一个等号右侧等于整题最终答案（允许先化简括号内再除等合法中间式）。
    """
    if is_algebraic_work_context(expr, student_work):
        return False
    expr_value = safe_eval_math_expr(expr)
    steps = split_equation_steps(student_work)
    if expr_value is None or len(steps) < 2:
        return False
    # 等号链每步成立且末步等于整题值：允许首行跳步（简便运算），不按「第一步未抄全原式」判错
    if not detect_invalid_equal_steps(student_work, expr, student_work):
        final_val = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
        if final_val is not None and values_equal(final_val, expr_value):
            return False
    exn = normalize_math_expr(re.sub(r"\s+", "", str(expr)))
    for i in range(len(steps) - 1):
        left_value = safe_eval_math_expr(steps[i])
        right_value = safe_eval_math_expr(steps[i + 1])
        if left_value is None or right_value is None:
            continue
        if not values_equal(left_value, right_value):
            if i != 0:
                return False
            s0 = normalize_math_expr(re.sub(r"\s+", "", steps[i]))
            if s0 == exn:
                return True
            # 左侧数值已达整题标准值，多为中间步骤正确、仅后续抄写/末步算错，不按「第一步从原式错」论处
            if values_equal(left_value, expr_value):
                return False
            return True
    first_left_value = safe_eval_math_expr(steps[0])
    if first_left_value is not None and not values_equal(first_left_value, expr_value):
        # 与 student_solution 一致：多步加减允许微小合并/取整差，不因跳步合并与冗长通分式略有出入判「第一步错」
        if _expr_allows_numeric_slack(expr) and values_equal(first_left_value, expr_value, tolerance=_MULTISTEP_ADDSUB_SLACK):
            return False
        return True
    return False


def looks_like_only_copied_question(expr: str, student_work: str) -> bool:
    """作答仅重复题干/原式，无有效变形或计算（与「未写」区分：有字但未算）。"""
    if not expr or not student_work:
        return False
    if is_algebraic_work_context(expr, student_work):
        return False
    raw = strip_invalidated_work(str(student_work)).strip()
    if not raw:
        return False
    ex = normalize_math_expr(re.sub(r"\s+", "", str(expr)))
    wx = normalize_math_expr(re.sub(r"\s+", "", raw.replace("\n", "")))
    if not ex or not wx:
        return False
    if "=" not in raw:
        return wx == ex or (ex in wx and len(wx) - len(ex) <= 2)
    parts = split_equation_steps(raw)
    if len(parts) < 2:
        seg = normalize_math_expr(parts[0].replace(" ", "")) if parts else wx
        return bool(seg) and seg == ex
    p0 = normalize_math_expr(parts[0].replace(" ", ""))
    p1 = normalize_math_expr(parts[1].replace(" ", ""))
    if p0 == ex and p1 == ex:
        return True
    if p0 == ex:
        v1 = safe_eval_math_expr(parts[1])
        ve = safe_eval_math_expr(expr)
        if v1 is not None and ve is not None and values_equal(v1, ve) and parts[0].strip() == parts[1].strip():
            return True
    return False


def looks_like_copied_other_question(expr: str, student_work: str) -> bool:
    """粗略识别把相邻题过程串到本题的情况。"""
    if not expr or not student_work:
        return False
    expr_norm = normalize_math_expr(expr)
    work_norm = normalize_math_expr(student_work)
    steps = split_equation_steps(work_norm)
    first = steps[0] if steps else work_norm
    expr_nums = set(re.findall(r"\d+", expr_norm))
    first_nums = set(re.findall(r"\d+", first))
    expr_ops = set(re.findall(r"[+\-*/]", expr_norm))
    first_ops = set(re.findall(r"[+\-*/]", first))
    if len(expr_nums) >= 2 and first_nums and expr_nums.isdisjoint(first_nums):
        return True
    if expr_ops and first_ops and expr_ops.isdisjoint(first_ops) and len(expr_nums & first_nums) <= 1:
        return True
    return False


def remove_invalidated_reason_fragments(reason: str) -> str:
    """从评语中移除把涂改/划掉内容当错误讲解的片段。"""
    if not reason:
        return ""
    text = str(reason)
    text = re.sub(r"[^。！？；;]*?(?:划掉|划去|涂掉|涂改|擦掉|作废|错写|废弃|被划掉|被划去|修改前|旧过程)[^。！？；;]*?[。！？；;]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def strip_false_negative_grading_hints(reason: str) -> str:
    """去掉与本地验算结论矛盾的「误错」套话（如跳步却判第一步整题错、结果实际一致却称末步不一致）。"""
    if not reason:
        return ""
    t = str(reason)
    t = re.sub(r"\s*提示：第一步已从原式错误变形[^。]*。", "", t)
    t = re.sub(r"[（(]?\s*订正提示[：:][^。\n]*?末步[^。\n]*?标准答案[^。\n]*", "", t)
    t = re.sub(r"订正提示：[^。]*?末步结果与标准答案不一致[^。]*。", "", t)
    t = re.sub(r"说明：题目数值结果应为[^。]*?与当前作答不一致。", "", t)
    t = re.sub(r"说明：按题干重算标准结果约为[^。]*?与作答过程或最终结果不一致。", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def student_solution_arithmetically_correct(expr: str, student_work: str, student_answer: str = "", expected_answer: str = "") -> bool:
    """本地确认横式/脱式是否每步成立且最终值正确。"""
    if looks_like_copied_other_question(expr, student_work) or looks_like_only_copied_question(expr, student_work):
        return False

    if is_symbolic_expression(expr):
        if is_algebraic_work_context(expr, student_work) and algebra_numeric_outcome_ok(
            expr, student_work, student_answer, expected_answer
        ):
            return True
        return False

    standard_value = safe_eval_math_expr(expr) or extract_math_value(expected_answer)
    if standard_value is None:
        return False
    if detect_invalid_equal_steps(student_work, expr, student_work):
        return False
    steps = split_equation_steps(student_work)
    if steps:
        first_value = safe_eval_math_expr(steps[0])
        final_value = safe_eval_math_expr(steps[-1]) or extract_math_value(steps[-1])
        if first_value is not None and not values_equal(first_value, standard_value):
            if not (_expr_allows_numeric_slack(expr) and values_equal(first_value, standard_value, tolerance=_MULTISTEP_ADDSUB_SLACK)):
                return False
        if final_value is not None:
            if values_equal(final_value, standard_value):
                return True
            if _expr_allows_numeric_slack(expr) and values_equal(final_value, standard_value, tolerance=_MULTISTEP_ADDSUB_SLACK):
                return True
            return False
    answer_value = extract_math_value(student_answer) or extract_math_value(student_work)
    if answer_value is None:
        return False
    if values_equal(answer_value, standard_value):
        return True
    if _expr_allows_numeric_slack(expr) and values_equal(answer_value, standard_value, tolerance=_MULTISTEP_ADDSUB_SLACK):
        return True
    return False


def local_arithmetic_guard(q: dict, status: str, process_score: float, result_score: float, structure_score: float, reason: str):
    """对 AI 结果做本地算术兜底，防止明显算式误判为正确。"""
    expr = q.get("expr", "")
    student_work = q.get("student_work", "")
    student_answer = q.get("student_answer", "")
    expected_answer = q.get("expected_answer", "")

    expr_value = safe_eval_math_expr(expr)
    expected_value = extract_math_value(expected_answer)
    standard_value = expr_value if expr_value is not None else expected_value
    answer_value = extract_math_value(student_answer) or extract_math_value(student_work)
    arith_chain_ok = student_solution_arithmetically_correct(
        expr, student_work, student_answer, q.get("expected_answer", "") or expected_answer
    )
    invalid_steps = detect_invalid_equal_steps(student_work, expr, student_work)
    copied_other_question = looks_like_copied_other_question(expr, student_work)
    first_step_wrong = first_step_invalid(expr, student_work)
    correct_steps = build_correct_steps(expr)
    symbolic = is_symbolic_expression(expr) or is_algebraic_work_context(expr, student_work)

    if copied_other_question:
        status = "未作答"
        process_score = 0
        result_score = 0
        structure_score = 0
        q["student_work"] = ""
        q["student_answer"] = ""
        reason = (f"本道题作答区域疑似为空，识别到的过程与题目数字/运算符不匹配，可能属于相邻题，"
                  f"不作为本题有效作答。{correct_steps}").strip()
        return status, process_score, result_score, structure_score, reason

    if looks_like_only_copied_question(expr, student_work):
        status = "错误"
        process_score = 0
        result_score = 0
        structure_score = 0
        reason = (
            f"{reason} 说明：作答仅重复题干或原式，未见「=」后的有效变形或计算，过程分、结果分、规范分均为 0。"
            f" {correct_steps}"
        ).strip()
        return status, process_score, result_score, structure_score, reason

    # 模型若已判「正确」但数值/等号链与标准不符，不得在此提前返回，否则末步算错也会被放过。

    if not symbolic and standard_value is not None:
        q["expected_answer"] = q.get("expected_answer") or fraction_to_text(standard_value)
        if answer_value is not None and not values_equal(answer_value, standard_value):
            if not arith_chain_ok:
                status = "错误"
                result_score = 0
                process_score = min(process_score, 0 if first_step_wrong else 5)
                reason = f"{reason} 说明：题目数值结果应为 {q['expected_answer']}，与当前作答不一致。{correct_steps}".strip()

    if not symbolic and invalid_steps:
        left, right, left_value, right_value = invalid_steps[0]
        status = "错误"
        process_score = 0 if first_step_wrong else min(process_score, 3)
        if standard_value is not None and answer_value is not None and values_equal(answer_value, standard_value):
            result_score = min(result_score, 2)
        reason = (f"{reason} 说明：过程等号不成立，{left}={fraction_to_text(left_value)}，但 {right}={fraction_to_text(right_value)}，"
                  f"不能判为正确。{correct_steps}").strip()

    if status == "错误" and not reason:
        reason = f"作答存在错误。{correct_steps}"
    elif status == "错误" and "正确步骤" not in reason:
        reason = f"{reason} {correct_steps}".strip()

    if status == "正确" and not symbolic:
        if invalid_steps:
            status = "错误"
            result_score = 0
            process_score = 0 if first_step_wrong else min(process_score, 3)
            left, right, left_value, right_value = invalid_steps[0]
            reason = (f"{reason} 说明：过程等号不成立，{left}={fraction_to_text(left_value)}，但 {right}={fraction_to_text(right_value)}，"
                      f"不能判为正确。{correct_steps}").strip()
        elif standard_value is not None and answer_value is not None and not values_equal(answer_value, standard_value):
            status = "错误"
            result_score = 0
            process_score = min(process_score, 0 if first_step_wrong else 5)
            q["expected_answer"] = q.get("expected_answer") or fraction_to_text(standard_value)
            reason = f"{reason} 说明：题目数值结果应为 {q['expected_answer']}，与当前作答不一致。{correct_steps}".strip()
        elif standard_value is not None and student_work and not arith_chain_ok:
            status = "错误"
            result_score = 0
            process_score = min(process_score, 0 if first_step_wrong else 5)
            q["expected_answer"] = q.get("expected_answer") or fraction_to_text(standard_value)
            reason = (
                f"{reason} 说明：按题干重算标准结果约为 {q['expected_answer']}，与作答过程或最终结果不一致。{correct_steps}"
            ).strip()

    return status, process_score, result_score, structure_score, reason


def normalize_question(q: dict) -> dict:
    """兜底规范 AI 返回，避免缺字段或误把未作答/过程错结果对判成正确。"""
    q = q or {}
    status = str(q.get("status", "错误")).strip()
    student_work_raw = strip_invalidated_work(str(q.get("student_work", q.get("work", "")) or "").strip())
    student_answer_raw = strip_invalidated_work(str(q.get("student_answer", q.get("answer", "")) or "").strip())
    expr = str(q.get("expr", "")).strip()
    expr = maybe_fix_expr_from_student_numeric(expr, student_work_raw, student_answer_raw)
    q["expr"] = expr
    student_work = student_work_raw
    student_answer = student_answer_raw
    reason = remove_invalidated_reason_fragments(str(q.get("reason", "")).strip())
    symbolic_expr = is_symbolic_expression(expr) or is_algebraic_work_context(expr, student_work)
    if expr and not is_symbolic_expression(expr):
        truth = safe_eval_math_expr(expr)
        exp_raw = str(q.get("expected_answer") or "").strip()
        if truth is not None and exp_raw:
            exp_v = extract_math_value(exp_raw)
            if exp_v is not None and not values_equal(exp_v, truth):
                q["expected_answer"] = fraction_to_text(truth)
    arithmetically_correct = student_solution_arithmetically_correct(expr, student_work, student_answer, q.get("expected_answer", ""))
    contradiction_correct = (
        status == "错误"
        and arithmetically_correct
        and not looks_like_only_copied_question(expr, student_work)
        and not first_step_invalid(expr, student_work)
    )
    if contradiction_correct:
        status = "正确"
        q["process_score"] = max(clamp_score(q.get("process_score"), 0, 6), 6)
        q["result_score"] = max(clamp_score(q.get("result_score"), 0, 3), 3)
        q["structure_score"] = max(clamp_score(q.get("structure_score"), 0, 1), 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    no_answer_markers = ("未作答", "空白", "没写", "无作答", "未填写", "没有作答", "答题区为空", "没有有效作答")
    blank_like_answers = ("", "无", "空", "空白", "未写", "未作答", "没写", "无作答", "没有")
    has_no_effective_answer = student_work in blank_like_answers and student_answer in blank_like_answers
    if status == "未作答" or (has_no_effective_answer and (not reason or any(m in reason for m in no_answer_markers))):
        status = "未作答"
        q["student_work"] = ""
        q["student_answer"] = ""
        q["process_score"] = 0
        q["result_score"] = 0
        q["structure_score"] = 0
        q["total_score"] = 0
        q["reason"] = sanitize_reason_text(reason or "题目有作答区域但没有看到有效书写，按未作答处理。")
        return q

    process_score = clamp_score(q.get("process_score"), 0, 6)
    result_score = clamp_score(q.get("result_score"), 0, 3)
    structure_score = clamp_score(q.get("structure_score"), 0, 1)

    error_keywords = (
        "过程错", "过程错误", "中间错", "等号不成立", "抄错", "列式错", "符号错", "括号",
        "进位错", "退位错", "借位错", "通分错", "约分错", "漏项", "单位错", "计算错",
        "94×20", "1880", "碰巧", "明显错误", "严重错误"
    )
    if status == "过程不规范" and any(k in reason for k in error_keywords) and not arithmetically_correct and not symbolic_expr:
        status = "错误"
        process_score = min(process_score, 3)
        if not reason.endswith("按错误处理。"):
            reason = f"{reason} 过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。"

    status, process_score, result_score, structure_score, reason = local_arithmetic_guard(
        q, status, process_score, result_score, structure_score, reason
    )

    arith_after_guard = student_solution_arithmetically_correct(
        expr, student_work, student_answer, q.get("expected_answer", "")
    )

    if (
        arith_after_guard
        and status == "错误"
        and not detect_invalid_equal_steps(student_work, expr, student_work)
        and not looks_like_only_copied_question(expr, student_work)
        and not first_step_invalid(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    if (
        status == "错误"
        and is_symbolic_expression(expr)
        and algebra_numeric_outcome_ok(expr, student_work, student_answer, q.get("expected_answer", ""))
        and not first_step_invalid(expr, student_work)
        and not looks_like_only_copied_question(expr, student_work)
    ):
        status = "正确"
        process_score = max(process_score, 6)
        result_score = max(result_score, 3)
        structure_score = max(structure_score, 1)
        reason = re.sub(r"\s*过程存在实质性错误，即使结果接近或碰巧正确，也按错误处理。", "", reason).strip()
        reason = re.sub(r"\s*正确步骤：.*?答案：[^。]*。", "", reason).strip()
        reason = strip_false_negative_grading_hints(reason) or "过程、每步等式和最终结果均正确。"

    if (
        status != "未作答"
        and student_work
        and first_step_invalid(expr, student_work)
        and str(status).strip() != "正确"
        and not arith_after_guard
    ):
        if opposite_sign_final_only(expr, student_answer, student_work):
            status = "错误"
            process_score = max(process_score, 4)
            result_score = 0
            reason = re.sub(r"\s*提示：第一步已从原式错误变形[^。]*。", "", reason).strip()
        else:
            status = "错误"
            process_score = 0
            result_score = 0
            structure_score = 0
            if "第一步" not in reason and "从原式错误变形" not in reason:
                reason = (reason + " 提示：第一步已从原式错误变形，后续步骤不再计过程分或结果分。").strip()

    if status == "错误" and result_score == 0 and process_score >= 5:
        reason_text = reason or "过程基本正确但最终答案错误。"
        if any(k in reason_text for k in ("结果", "答案", "最终", "计算细节")) and not any(k in reason_text for k in error_keywords):
            process_score = min(process_score, 6)

    # 方程/配方类：仅书写格式、未写解集/检验、下标写法等，不单独构成「过程不规范」
    if _equation_or_algebra_hint(expr, student_work) and status == "过程不规范":
        if not any(k in reason for k in error_keywords) and result_score >= 2 and process_score >= 4:
            status = "正确"
            process_score = max(process_score, 6)
            result_score = max(result_score, 3)
            structure_score = max(structure_score, 1)

    calculated_total = process_score + result_score + structure_score
    total_score = clamp_score(calculated_total, 0, 10)
    if status == "正确" and total_score < 9:
        status = "过程不规范" if total_score >= 7 else "错误"
    if status == "错误" and total_score > 8:
        total_score = 8
    if status in ("错误", "未作答") and expr and "正确步骤" not in reason and not symbolic_expr:
        reason = f"{reason} {build_correct_steps(expr)}".strip()

    if str(status).strip() == "正确":
        reason = strip_false_negative_grading_hints(reason)

    q["expr"] = expr
    q["student_work"] = student_work
    q["student_answer"] = student_answer
    q["status"] = status if status in ("正确", "过程不规范", "错误", "未作答") else "错误"
    q["process_score"] = process_score
    q["result_score"] = result_score
    q["structure_score"] = structure_score
    q["total_score"] = clamp_score(total_score, 0, 10)
    q["reason"] = sanitize_reason_text(reason or "已按题目、过程和最终答案综合核验。")
    return q


def normalize_math_result(data: dict) -> dict:
    if not data:
        return data
    questions = [normalize_question(q) for q in data.get("questions", [])]
    data["questions"] = questions
    statuses = [q.get("status") for q in questions]
    if not questions or all(s == "未作答" for s in statuses):
        data["overall_result"] = "未作答"
    elif all(s == "正确" for s in statuses):
        data["overall_result"] = "全部正确"
    elif all(s in ("错误", "未作答") for s in statuses):
        data["overall_result"] = "全部错误"
    else:
        data["overall_result"] = "部分正确"
    return data


def _teacher_context_suffix_math(grade_level: str, teacher_note: str) -> str:
    g = (grade_level or "").strip()
    n = (teacher_note or "").strip()
    if not g and not n:
        return ""
    parts = ["\n\n【任课教师补充说明（请纳入批改语境；若与卷面或题干明显冲突以卷面与题干为准）】"]
    if g:
        parts.append(
            f"学生所在年级（老师填写）：{g}。请按该学段课标与常见认知水平把握判分尺度、总评语气与 reason 详略；不要超纲使用明显超出该学段要求的术语硬套学生。"
        )
    if n:
        parts.append(
            f"老师备注 / 学情说明：{n}\n请结合以上内容辅助判断掌握程度与常见问题，并在总评、薄弱点或书写习惯建议中酌情呼应（无需重复摘抄备注全文）。"
        )
    return "\n".join(parts)


def call_ai_math(image_path: str, *, grade_level: str = "", teacher_note: str = "") -> dict:
    try:
        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        image_uri = f"data:{get_mime_type(image_path)};base64,{b64}"

        suffix = _teacher_context_suffix_math(grade_level, teacher_note)
        messages = [{"role": "user", "content": [
            {"image": image_uri},
            {"text": MATH_PROMPT + suffix}
        ]}]

        # max_length 提高可减少「只返回前几题 JSON」被截断的情况（长卷多题须完整输出 questions）
        resp = dashscope.MultiModalConversation.call(
            model="qwen-vl-max",
            messages=messages,
            timeout=120,
            max_length=8192,
            temperature=0.05,
            top_p=0.3,
        )
        if resp.status_code == HTTPStatus.OK:
            raw = resp.output.choices[0].message.content[0]["text"]
            return normalize_math_result(json.loads(extract_json_safe(raw)))
        logging.error(f"API 失败：{resp.status_code} | {resp.message}")
        return None
    except Exception as e:
        logging.error(f"处理异常：{e}")
        return None


# ✅ 修复点 1：参数名改为 data，if 补全 data
def format_console_output(data: dict, filename: str):
    if not data:
        print(f"{Fore.RED}❌ 批改失败\n")
        return

    status = data.get("overall_result", "未知")
    status_map = {
        "全部正确": (Fore.GREEN, "✅ 全部正确"),
        "部分正确": (Fore.YELLOW, "⚠️ 部分正确"),
        "全部错误": (Fore.RED, "❌ 全部错误"),
        "未作答": (Fore.BLUE, "📝 未作答")
    }
    color, status_text = status_map.get(status, (Fore.WHITE, status))

    print(f"\n{Fore.CYAN}{'=' * 70}")
    print(f"📄 文件：{filename}")
    print(f"🏷️  状态：{color}{status_text}{Style.RESET_ALL}")

    print(f"\n📝 逐题详情:")
    for q in data.get("questions", []):
        q_status = q.get("status", "")
        q_color = Fore.GREEN if q_status == "正确" else Fore.YELLOW if q_status == "过程不规范" else Fore.RED if q_status == "错误" else Fore.BLUE
        score_text = f"{q.get('total_score', 0):g}/10"
        print(
            f"  {q['id']}. {q_color}{prettify_math(q.get('expr', ''))}{Style.RESET_ALL} → {q_color}{q_status}{Style.RESET_ALL} ({score_text})")
        if q.get("student_work"):
            print(f"     {Style.DIM}学生过程：{prettify_math(q['student_work'])}{Style.RESET_ALL}")
        if q.get("expected_answer"):
            print(f"     {Style.DIM}标准结果：{prettify_math(q['expected_answer'])}{Style.RESET_ALL}")
        print(
            f"     {Style.DIM}得分：过程 {q.get('process_score', 0):g}/6，结果 {q.get('result_score', 0):g}/3，规范 {q.get('structure_score', 0):g}/1{Style.RESET_ALL}")
        print(f"     {Style.DIM}└─ {q.get('reason', '')}{Style.RESET_ALL}")

    print(f"\n💬 老师评语:")
    for line in textwrap.wrap(data.get("personal_comment", "暂无"), width=66, initial_indent="   ",
                              subsequent_indent="   "):
        print(line)

    print(f"\n⚠️  薄弱知识点:")
    weaks = data.get("weak_points", [])
    for w in (weaks if weaks else ["暂无显著薄弱点，基础扎实，继续保持！"]):
        print(f"   • {w}")
    print(f"{Fore.CYAN}{'=' * 70}{Style.RESET_ALL}\n")


def generate_class_analysis(results: list) -> str:
    if not results: return "无有效数据"
    weak_counter, status_counter = {}, {"全部正确": 0, "部分正确": 0, "全部错误": 0, "未作答": 0}
    for r in results:
        status_counter[r.get("overall_result", "全部错误")] += 1
        for pt in r.get("weak_points", []): weak_counter[pt] = weak_counter.get(pt, 0) + 1

    top_weak = sorted(weak_counter.items(), key=lambda x: x[1], reverse=True)[:5]
    lines = ["========================================", "📊 数学作业学情分析报告",
             "========================================",
             f"📝 总份数：{len(results)}", f"✅ 全部正确：{status_counter['全部正确']}",
             f"⚠️ 部分正确：{status_counter['部分正确']}",
             f"❌ 全部错误：{status_counter['全部错误']}", f"📝 未作答：{status_counter['未作答']}", "",
             "🔥 高频薄弱知识点 (Top 5):"]
    for k, v in top_weak: lines.append(f"   - {k}: {v} 人次")
    lines.append("\n💡 教学建议:")
    lines.append(f"   建议针对【{top_weak[0][0]}】开展专项训练。" if top_weak else "   整体掌握良好，继续保持。")
    return "\n".join(lines)


# ================= 5. 主流程 =================
def batch_correct_math():
    folder = Path(MATH_FOLDER)
    if not folder.exists():
        print(f"{Fore.RED}❌ 请创建 {MATH_FOLDER} 文件夹");
        return

    imgs = sorted(list(folder.glob("*.jpg")) + list(folder.glob("*.png")) + list(folder.glob("*.jpeg")))
    if not imgs:
        print(f"{Fore.YELLOW}⚠️ 文件夹内无图片");
        return

    print(f"{Fore.CYAN}📚 启动 AI 数学作业批改系统...\n")
    results = []

    for img in imgs:
        print(f"{Fore.YELLOW}⏳ 批改：{img.name}")
        data = call_ai_math(img)
        # ✅ 修复点 2：if 后面补全 data
        if data:
            data["filename"] = img.name
            results.append(data)
            format_console_output(data, img.name)
        else:
            print(f"{Fore.RED}   ❌ 失败\n")

    if not results:
        print(f"{Fore.RED}\n⚠️ 未成功批改任何作业");
        return

    print("\n" + generate_class_analysis(results))

    wb = Workbook()
    ws = wb.active;
    ws.title = "数学批改"
    headers = ["文件名", "总评", "平均分", "逐题详情 (JSON)", "评语", "薄弱点"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True);
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        questions = r.get("questions", [])
        avg_score = round(sum(q.get("total_score", 0) for q in questions) / len(questions), 1) if questions else 0
        ws.append([r["filename"], r["overall_result"], avg_score,
                   json.dumps(questions, ensure_ascii=False),
                   r.get("personal_comment", ""), ", ".join(r.get("weak_points", []))])
    ws.column_dimensions['D'].width = 80
    wb.save(OUTPUT_EXCEL)

    with open(OUTPUT_TXT, "w", encoding="utf-8") as f:
        f.write(generate_class_analysis(results) + "\n\n")
        for r in results:
            f.write(f"--- {r['filename']} | {r['overall_result']} ---\n")
            for q in r.get("questions", []):
                f.write(f"  {q['id']}. {prettify_math(q['expr'])} → {q['status']} ({q['reason']})\n")
            f.write(f"  评语：{r.get('personal_comment', '')}\n\n")

    print(f"\n{Fore.GREEN}✅ 完成！报告已保存至 {OUTPUT_EXCEL} 和 {OUTPUT_TXT}")


if __name__ == "__main__":
    batch_correct_math()